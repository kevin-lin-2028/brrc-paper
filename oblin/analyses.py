
from __future__ import annotations

import json
import math
import multiprocessing as mp
import random
import re
import shutil
import sys
import time
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

import RNA
import numpy as np
import openpyxl
import requests
from Bio import Phylo
from scipy.stats import mannwhitneyu, norm

from oblin.core import *
from oblin.core import JSON, PROC, RAW, ROOT

HELP = "controls: positive | bacterial | matched | circular | snapshot"

OB_LEN_LO, OB_LEN_HI = 700, 1700

RFAM_FAMS = [
    "RF00010.fa", "RF00011.fa", "RF00012.fa", "RF00013.fa", "RF00023.fa",
    "RF00028.fa", "RF00029.fa", "RF00031.fa", "RF00050.fa", "RF00059.fa",
    "RF00162.fa", "RF00174.fa", "RF00177.fa", "RF00379.fa", "RF01051.fa",
    "RF01057.fa", "RF01998.fa",
]

VIROIDS = {
    "PSTVd_NC_002030": (
        "CGGAACTAAACTCGTGGTTCCTGTGGTTCACACCTGACCTCCTGAGCAGAAAAGAAAAAA"
        "GAAGGCGGCTCGGAGGAGCGCTTCAGGGATCCCCGGGGAAACCTGGAGCGAACTGGCAAA"
        "AAAGGACGGTGGGGAGTGCCCAGCGGCCGACAGGAGTAATTCCCGCCGAAACAGGGTTTT"
        "CACCCTTCCTTTCTTCGGGTGTCCTTCCTCGCGCCCGCAGGACCACCCCTCGCCCCCTTT"
        "GCGCTGTCGCTTCGGCTACTACCCGGTGGAAACAACTGAAGCTCCCGAGAACCGCTTTTT"
        "CTCTATCTTACTTGCTTCGGGGCGAGGGTGTTTAGCCCTTGGAACCGCAGTTGGTTCCT"
    ),
    "ASBVd_NC_001340": (
        "TTCAGGGATCCCAACTACATTGCTGTTGCAGTGCAGCGTGTGTAGTTGGGAATTGCTACA"
        "TGCATTCATGGAACATTCAGGGTGCAACTGAGTGGTGCGAACATGCATCAGGAACTGTGC"
        "ACGGCAGTTCATGCGCAACTAGAACAACTAGTGCATGGTGCAACAAATTCAGTGCAATCC"
        "AGTGCATTGTAGCAGTGAACAACTGCGCATGCATGCATTGAACTGTGCAACTGCAATCCA"
        "AGTGCATGCATGCAGCAATTCAGTGCATTC"
    ),
    "HSVd_NC_001351": (
        "GGGCAACTCTTCTCAGAATCCAGCGAGAGGCAAACAAGTGAAAACAATTACCAACAAAGA"
        "AAACAGAAAGCAACGGAGAACAGTGAACATGGGTAATCCAGAGTGTTGGGAATGAATTGT"
        "CCCATCCCAGCTACTTTCTGTGGTTCACACCTGACCTCCTGAGCAGAAAAGAAAAAGAAG"
        "GCGGCTCGGAGGAGCGCTTCAGGGATCCCCGGGGAAACCTGGAGCGAACTGGCAAAAAAG"
        "GACGGTGGGGAGTGCCCAGCGGCCGACAGGAGTAATTCCCGCCGAAACAGGG"
    ),
    "CChMVd_NC_003540": (
        "AGGGAUCCUGCUUCAGCUUUCCAGUGGCAGCAUUUUUCCAGCUUCAGGAGUGGGCUUUCC"
        "UCAGAGUUUUUUUCCUUCUCCAGCAGGAUUCCUUCUCCAGCAGGAUUCCUUCUUCAGCAU"
        "UCCUUCUUCAGCAUUCCUUCUUCAGCAUUCCUUCUUCAGCAUUCCUUCUUCAGCAUUCCU"
        "UCUUCAGCAUUCCUUCUUCAGCAUUCCUUCUUCAGCAUUCCUUCUUCAGCAUUCCUUCUU"
        "CAGCAUUCCUUCUUCAGCAUUCCUUCUCCAGCAGGAUCCCU"
    ),
    "CCCVd_NC_001464": (
        "CCAGGGGAAACCTGGAGCGAACTGGCAAAAAAGGACGGTGGGGAGTGCCCAGCGGCCGAC"
        "AGGAGTAATTCCCGCCGAAACAGGGTTTTCACCCTTCCTTTCTTCGGGTGTCCTTCCTCG"
        "CGCCCGCAGGACCACCCCTCGCCCCCTTTGCGCTGTCGCTTCGGCTACTACCCGGTGGAA"
        "ACAACTGAAGCTCCCGAGAACCGCTTTTTCTCTATCTTACTTGCTTCGGGGCGAGGGTGT"
        "TTAGCCCTTGGAACCGCAGTTGGT"
    ),
    "TASVd_NC_001553": (
        "CGGAACTAAACTCGTGGTTCCTGTGGTTCACACCTGACCTCCTGAGCAGAAAAGAAAAAG"
        "AAAACCAGTTTTGCAGCATCAGCAGGAGCAGTGGAGCGCAGGGAATTCAGCAGTTGCCAA"
        "AAAAGGACGGTGGGGAGTGCCCAGCGGCCGACAGGAGTAATTCCCGCCGAAACAGGGTTT"
        "TCACCCTTCCTTTCTTCGGGTGTCCTTCCTCGCGCCCGCAGGACCACCCCTCGCCCCCTT"
        "TGCGCTGTCGCTTCGGCTACTACCCGGTGGAAACAACTGAAGCTCCCGAGAACCGCTTTT"
        "TCTCTATCTTACTTGCT"
    ),
    "CSVd_NC_002015": (
        "CGGAACTAAACTCGTGGTTCCTGTGGTTCACACCTGACCTCCTGAGCAGAAAAGAAAAAA"
        "GAAGGCGGCTCGGAGGAGCGCTTCAGGGATCCCCGGGGAAACCTGGAGCGAACTGGCAAA"
        "AAAGGACGGTGGGGAGTGCCCAGCGGCCGACAGGAGTAATTCCCGCCGAAACAGGGTTTT"
        "CACCCTTCCTTTCTTCGGGTGTCCTTCCTCGCGCCCGCAGGACCACCCCTCGCCCCCTTT"
        "GCGCTGTCGCTTCGGCTACTACCCGGTGGAAACAACTGAAGCTCCCGAGAACCGCTTTTT"
        "CTCTATCTTACTTGCTTCGGG"
    ),
}

def _fasta(path):
    out = []
    nm, buf = None, []
    with open(path) as fh:
        for line in fh:
            line = line.rstrip()
            if line.startswith(">"):
                if nm is not None:
                    out.append((nm, "".join(buf)))
                nm = line[1:]; buf = []
            elif line:
                buf.append(line)
        if nm is not None:
            out.append((nm, "".join(buf)))
    return out

def _controls_gc(s):
    s = s.upper()
    n = sum(1 for c in s if c in "GC")
    L = sum(1 for c in s if c in "ACGUT")
    return n / max(1, L)

def _topo(db):
    n = len(db)
    paired = sum(1 for c in db if c in "()")
    fp = paired / n if n else 0.0
    loops, cur = [], 0
    for c in db:
        if c == ".":
            cur += 1
        else:
            if cur > 0:
                loops.append(cur)
            cur = 0
    if cur > 0:
        loops.append(cur)
    sld = 100.0 * sum(1 for L in loops if L <= 5) / n if n else 0.0
    ml = max(loops) if loops else 0
    return round(fp, 4), round(sld, 2), int(ml)

def _passes(fp, sld, ml):
    return fp > 0.65 and sld > 10 and ml <= 15

def _circ_fold(seq):
    s = seq.upper().replace("T", "U")
    md = RNA.md(); md.circ = 1
    fc = RNA.fold_compound(s, md)
    return fc.mfe()

def positive():
    print(f"viroid panel ({len(VIROIDS)} sequences)", flush=True)
    vir_rows = []
    for nm, seq in VIROIDS.items():
        db, mfe = _circ_fold(seq)
        fp, sld, ml = _topo(db)
        vir_rows.append({"id": nm, "length": len(seq),
                         "frac_paired": fp, "sld_per100nt": sld,
                         "max_loop": ml, "brrc_pass": _passes(fp, sld, ml),
                         "mfe_kcal_per_mol": round(mfe, 2)})
        flag = "PASS" if vir_rows[-1]["brrc_pass"] else "fail"
        print(f"  {nm:24s} {len(seq):4d} nt  fp={fp:.3f} sld={sld:5.2f} "
              f"max={ml:3d}  -> {flag}", flush=True)
    n_vir_pass = sum(1 for r in vir_rows if r["brrc_pass"])
    print(f"  viroids: {n_vir_pass}/{len(vir_rows)} pass", flush=True)

    print("\nHDV / delta-like genomes", flush=True)
    delta_seqs = dict(_fasta(RAW / "delta_full_genomes.fasta"))
    print(f"  loaded {len(delta_seqs)}", flush=True)
    delta_rows = []
    for i, (nm, seq) in enumerate(delta_seqs.items(), 1):
        try:
            db, mfe = _circ_fold(seq)
        except Exception as exc:
            delta_rows.append({"id": nm, "length": len(seq), "error": str(exc)})
            continue
        fp, sld, ml = _topo(db)
        delta_rows.append({"id": nm, "length": len(seq),
                           "frac_paired": fp, "sld_per100nt": sld,
                           "max_loop": ml, "brrc_pass": _passes(fp, sld, ml),
                           "mfe_kcal_per_mol": round(mfe, 2)})
        if i % 10 == 0:
            print(f"    folded {i}/{len(delta_seqs)}", flush=True)

    delta_ok = [r for r in delta_rows if "error" not in r]
    n_delta_pass = sum(1 for r in delta_ok if r["brrc_pass"])

    out = {
        "tool": "positive_controls",
        "viroids": {"n": len(vir_rows), "n_passing_brrc": n_vir_pass,
                    "pass_pct": round(100 * n_vir_pass / len(vir_rows), 2),
                    "rows": vir_rows},
        "delta_class": {"n_attempted": len(delta_rows),
                        "n_valid_folds": len(delta_ok),
                        "n_passing_brrc": n_delta_pass,
                        "pass_pct": (round(100 * n_delta_pass / len(delta_ok), 2)
                                     if delta_ok else None),
                        "rows": delta_rows},
        "comparator_summary": {
            "ob_circ_pass_pct": 88.95,
            "viroid_pass_pct": round(100 * n_vir_pass / len(vir_rows), 2),
            "delta_class_pass_pct": (round(100 * n_delta_pass / len(delta_ok), 2)
                                     if delta_ok else None),
        },
    }
    out_path = JSON / "positive_controls.json"
    out_path.write_text(json.dumps(out, indent=2))
    print(f"\nwrote {out_path}")
    return out

def bacterial():
    base = RAW / "rfam_controls"
    snapshot = RAW / "bacterial_brrc_window_sample.fasta"
    fam_n = Counter()
    cands = []
    if base.exists() and any(base.iterdir()):
        print(f"surveying {len(RFAM_FAMS)} bacterial Rfam families", flush=True)
        for f in RFAM_FAMS:
            p = base / f
            if not p.exists():
                continue
            for nm, s in _fasta(p):
                if OB_LEN_LO <= len(s) <= OB_LEN_HI:
                    cands.append((f.replace(".fa", ""), nm, s))
                    fam_n[f] += 1
        if len(cands) > 800:
            by_fam = defaultdict(list)
            for c in cands:
                by_fam[c[0]].append(c)
            cap = max(20, 800 // max(len(by_fam), 1))
            cands = sum((items[:cap] for items in by_fam.values()), [])[:800]
            print(f"  stratified subsample to {len(cands):,} (cap {cap}/family)",
                  flush=True)
    elif snapshot.exists():
        print(f"using snapshot {snapshot.name}", flush=True)
        for hdr, s in _fasta(snapshot):
            fam, _, orig = hdr.partition("|")
            cands.append((fam, orig, s))
            fam_n[fam + ".fa"] += 1
    else:
        raise FileNotFoundError(
            f"need {base} or {snapshot}; download from EBI Rfam FTP")

    print(f"\nfolding {len(cands):,} sequences", flush=True)
    rows, t0 = [], time.time()
    for i, (fam, nm, s) in enumerate(cands, 1):
        try:
            db, _ = _circ_fold(s)
        except Exception as exc:
            rows.append({"family": fam, "id": nm, "length": len(s),
                         "error": str(exc)})
            continue
        fp, sld, ml = _topo(db)
        rows.append({"family": fam, "id": nm, "length": len(s),
                     "gc": round(_controls_gc(s), 3),
                     "frac_paired": fp, "sld_per100nt": sld,
                     "max_loop": ml, "brrc_pass": _passes(fp, sld, ml)})
        if i % 100 == 0:
            r = i / (time.time() - t0)
            print(f"  {i}/{len(cands):,}  ({r:.1f}/s)", flush=True)

    ok = [r for r in rows if "error" not in r]
    npass = sum(1 for r in ok if r["brrc_pass"])
    by_n, by_pass = Counter(), Counter()
    for r in ok:
        by_n[r["family"]] += 1
        if r["brrc_pass"]:
            by_pass[r["family"]] += 1
    per_fam = {f: {"n": by_n[f], "n_pass": by_pass[f],
                   "pass_pct": round(100 * by_pass[f] / by_n[f], 2)}
               for f in by_n}

    out = {
        "tool": "bacterial_controls",
        "n_attempted": len(rows),
        "n_valid_folds": len(ok),
        "n_passing_brrc": npass,
        "pass_pct": round(100 * npass / max(len(ok), 1), 2),
        "length_window_nt": [OB_LEN_LO, OB_LEN_HI],
        "families_used": list(fam_n),
        "per_family": per_fam,
        "comparator_summary": {
            "ob_circ_pass_pct": 88.95,
            "HDV_delta_class_pass_pct": 100.00,
            "real_bacterial_RNAs_in_window_pass_pct": round(
                100 * npass / max(len(ok), 1), 2),
            "dinuc_shuf_pct": 0.30,
        },
    }
    p_out = JSON / "bacterial_controls.json"
    p_out.write_text(json.dumps(out, indent=2))
    print(f"\nwrote {p_out}")
    print(f"  pass: {npass}/{len(ok)} = {out['pass_pct']:.2f}%")
    print("  top 10 families by sample size:")
    for f, info in sorted(per_fam.items(), key=lambda kv: -kv[1]["n"])[:10]:
        print(f"    {f:8s}  n={info['n']:4d}  "
              f"pass={info['n_pass']:4d} ({info['pass_pct']:5.1f}%)")
    return out

def matched():
    obs = _fasta(PROC / "obelisks_zheludev_catalog.fasta")
    rfam = _fasta(RAW / "rfam_controls_expanded.fasta")
    print(f"loaded {len(obs):,} obelisks and {len(rfam):,} Rfam")

    L = np.array([len(s) for _, s in obs])
    G = np.array([_controls_gc(s) for _, s in obs])
    L_lo, L_hi = np.percentile(L, [5, 95])
    G_lo, G_hi = np.percentile(G, [5, 95])
    print(f"obelisk window: L=[{int(L_lo)},{int(L_hi)}] nt; "
          f"GC=[{G_lo:.2f},{G_hi:.2f}]")

    in_window = [(nm, s) for nm, s in rfam
                 if L_lo <= len(s) <= L_hi and G_lo <= _controls_gc(s) <= G_hi]
    in_lonly = [(nm, s) for nm, s in rfam if L_lo <= len(s) <= L_hi]
    print(f"  L+GC match: {len(in_window):,};  L-only: {len(in_lonly):,}")

    rng = np.random.default_rng(42)
    if len(in_window) > 200:
        in_window = [in_window[i] for i in rng.choice(len(in_window), 200,
                                                       replace=False)]
    if len(in_lonly) > 200:
        in_lonly = [in_lonly[i] for i in rng.choice(len(in_lonly), 200,
                                                     replace=False)]

    def fold_set(items, label):
        print(f"folding {len(items)} {label}")
        out = []
        for nm, s in items:
            su = s.upper().replace("T", "U")
            db, mfe = RNA.fold(su)
            fp, sld, ml = _topo(db)
            out.append({"name": nm.split()[0], "length": len(su),
                        "gc": _controls_gc(su), "mfe_per_nt": mfe / len(su),
                        "frac_paired": fp, "sld_per100nt": sld,
                        "max_loop": ml,
                        "passes_BRRC": _passes(fp, sld, ml)})
        return out

    rows = fold_set(in_window, "L+GC matched")
    rows_lo = fold_set(in_lonly, "L-only matched")

    fc = json.load(open(JSON / "catalog_features.json"))
    ob = fc["OBELISK_FULL"]
    fp_arr = np.array([r["frac_paired"] for r in rows])
    sld_arr = np.array([r["sld_per100nt"] for r in rows])
    ml_arr = np.array([r["max_loop"] for r in rows])
    pass_arr = np.array([r["passes_BRRC"] for r in rows])
    pass_lo = np.array([r["passes_BRRC"] for r in rows_lo])
    ob_pass = np.mean([_passes(r["frac_paired"],
                                r["sld_per100nt"],
                                r["max_loop"]) for r in ob])

    summary = {
        "tool": "matched_rfam_control",
        "match_window": {"length_nt_5pct": float(L_lo),
                          "length_nt_95pct": float(L_hi),
                          "GC_5pct": float(G_lo), "GC_95pct": float(G_hi)},
        "n_matched_rfam_length_and_GC": int(len(rows)),
        "n_rfam_len_only": int(len(rows_lo)),
        "n_matched_rfam_length_and_GC_folded": len(rows),
        "n_matched_rfam_len": len(rows_lo),
        "matched_rfam_length_and_GC_BRRC_pass_pct":
            float(100 * pass_arr.mean()),
        "matched_rfam_length_only_BRRC_pass_pct":
            float(100 * pass_lo.mean()) if len(pass_lo) else 0.0,
        "matched_rfam_BRRC_pass_pct": float(100 * pass_arr.mean()),
        "obelisk_BRRC_pass_pct": float(100 * ob_pass),
        "rfam_fp_mean": float(fp_arr.mean()),
        "matched_rfam_sld_mean": float(sld_arr.mean()),
        "rfam_ml_mean": float(ml_arr.mean()),
        "ob_fp_mean": float(np.mean(
            [r["frac_paired"] for r in ob])),
        "obelisk_sld_mean": float(np.mean(
            [r["sld_per100nt"] for r in ob])),
        "obelisk_max_loop_mean": float(np.mean(
            [r["max_loop"] for r in ob])),
        "matched_rfam_rows": rows,
        "rfam_len_only_rows": rows_lo,
    }
    out_path = JSON / "matched_rfam_control.json"
    out_path.write_text(json.dumps(summary, indent=2))
    print(f"\nwrote {out_path}")
    print(f"  L+GC matched pass: "
          f"{summary['matched_rfam_length_and_GC_BRRC_pass_pct']:.1f}%")
    print(f"  L-only matched pass: "
          f"{summary['matched_rfam_length_only_BRRC_pass_pct']:.1f}%")
    print(f"  obelisk catalog pass: {summary['obelisk_BRRC_pass_pct']:.1f}%")
    return summary

def _fold_one(args):
    nm, seq = args
    su = seq.upper().replace("T", "U")
    try:
        md = RNA.md(); md.circ = 1
        fc = RNA.fold_compound(su, md)
        db, mfe = fc.mfe()
    except Exception as exc:
        return {"id": nm.split()[0], "length": len(su), "error": str(exc)}
    fp, sld, ml = _topo(db)
    return {"id": nm.split()[0], "length": len(su),
            "mfe_per_nt": mfe / len(su),
            "frac_paired": fp, "sld_per100nt": sld,
            "max_loop": ml, "passes_BRRC": _passes(fp, sld, ml)}

def circular(limit=None):
    seqs = _fasta(PROC / "obelisks_zheludev_catalog.fasta")
    if limit:
        seqs = seqs[:int(limit)]
    n = len(seqs)
    workers = max(1, mp.cpu_count() - 1)
    print(f"folding {n:,} obelisks circularly on {workers} workers",
          flush=True)

    rows, t0 = [], time.time()
    with mp.Pool(workers) as pool:
        for i, r in enumerate(pool.imap_unordered(_fold_one, seqs,
                                                   chunksize=8), 1):
            rows.append(r)
            if i % 250 == 0:
                rate = i / (time.time() - t0)
                eta = (n - i) / rate if rate > 0 else 0
                print(f"  {i:,}/{n:,}  [{rate:.1f}/s, ETA {eta:.0f}s]",
                      flush=True)

    npass = sum(1 for r in rows if r.get("passes_BRRC"))
    n_fp = sum(1 for r in rows if r.get("frac_paired", 0) > 0.65)
    n_sld = sum(1 for r in rows
                if r.get("sld_per100nt", 0) > 10)
    n_ml = sum(1 for r in rows if r.get("max_loop", 99) <= 15)

    summary = {
        "tool": "circular",
        "n_obelisks_folded": len(rows),
        "circular_BRRC_pass_pct": round(100 * npass / len(rows), 2),
        "circ_fp_gt065_pct": round(100 * n_fp / len(rows), 2),
        "circular_sld_gt_10_pct": round(100 * n_sld / len(rows), 2),
        "circ_ml_le15_pct": round(100 * n_ml / len(rows), 2),
        "linear_reference_BRRC_pass_pct": 87.7,
    }
    out = JSON / "circular.json"
    out.write_text(json.dumps({"summary": summary, "rows": rows}, indent=2))
    print(f"\nwrote {out}")
    print(f"  circular BRRC pass: {summary['circular_BRRC_pass_pct']}% "
          f"(n={len(rows)}); linear ref: 87.7%")
    return summary

def snapshot():
    base = RAW / "rfam_controls"
    cands = []
    for f in RFAM_FAMS:
        p = base / f
        if not p.exists():
            continue
        for nm, s in _fasta(p):
            if OB_LEN_LO <= len(s) <= OB_LEN_HI:
                cands.append((f.replace(".fa", ""), nm, s))

    by_fam = defaultdict(list)
    for c in cands:
        by_fam[c[0]].append(c)
    cap = max(20, 800 // max(len(by_fam), 1))
    sampled = sum((items[:cap] for items in by_fam.values()), [])[:800]

    out = RAW / "bacterial_brrc_window_sample.fasta"
    with open(out, "w") as fh:
        for fam, nm, s in sampled:
            fh.write(f">{fam}|{nm}\n")
            for i in range(0, len(s), 60):
                fh.write(s[i:i+60] + "\n")
    print(f"wrote {out} ({len(sampled)} sequences, "
          f"{out.stat().st_size/1024:.1f} KB)")

CMDS = {
    "positive":  lambda argv: positive(),
    "bacterial": lambda argv: bacterial(),
    "matched":   lambda argv: matched(),
    "circular":  lambda argv: circular(argv[1] if len(argv) > 1 else None),
    "snapshot":  lambda argv: snapshot(),
}

def controls_main(argv=None):
    argv = sys.argv[1:] if argv is None else argv
    cmd = argv[0] if argv else "positive"
    if cmd not in CMDS:
        print(HELP, file=sys.stderr); sys.exit(2)
    CMDS[cmd](argv)

if __name__ == "__main__":
    controls_main()
def _rc(s):
    return s.translate(str.maketrans("ACGTUacgtu", "TGCAATGCAA"))[::-1]

def crispr(seed_k=20, max_mismatch=2):
    print("loading obelisks", flush=True)
    seqs = _fasta(PROC / "obelisks_zheludev_catalog.fasta")

    obelisks = [(nm, s.upper().replace("U", "T") + s.upper().replace("U", "T")[:50])
                for nm, s in seqs]
    print(f"  obelisks: {len(obelisks)}", flush=True)

    print("loading spacers", flush=True)
    spacers = _fasta(RAW / "spacer_seqName.fsa")
    print(f"  spacers: {len(spacers)}", flush=True)

    print(f"building spacer {seed_k}-mer seed index", flush=True)
    seed_index = defaultdict(list)
    n_seeds = 0
    for sp_idx, (sp_name, sp_seq) in enumerate(spacers):
        sp_seq = sp_seq.upper().replace("U", "T")
        if len(sp_seq) < seed_k:
            continue
        for strand_name, strand_seq in (("+", sp_seq), ("-", _rc(sp_seq))):
            for i in range(0, len(strand_seq) - seed_k + 1, max(1, seed_k // 2)):
                seed = strand_seq[i:i + seed_k]
                seed_index[seed].append((sp_idx, strand_name, i, strand_seq))
                n_seeds += 1
    print(f"  indexed {n_seeds} seeds from {len(spacers)} spacers", flush=True)

    print(f"scanning obelisks (seed match + mismatch extension; "
          f"max_mm={max_mismatch})", flush=True)
    hits = []
    t0 = time.time()
    for i, (nm, s) in enumerate(obelisks, 1):
        for pos in range(len(s) - seed_k + 1):
            seed = s[pos:pos + seed_k]
            if seed not in seed_index:
                continue
            for sp_idx, strand, sp_off, full_sp in seed_index[seed]:

                ob_start = pos - sp_off
                ob_end = ob_start + len(full_sp)
                if ob_start < 0 or ob_end > len(s):
                    continue
                ob_window = s[ob_start:ob_end]
                if len(ob_window) != len(full_sp):
                    continue
                mm = sum(1 for a, b in zip(ob_window, full_sp) if a != b)
                if mm <= max_mismatch:
                    hits.append({
                        "obelisk_id": nm,
                        "obelisk_pos": ob_start,
                        "spacer_idx": int(sp_idx),
                        "spacer_seq": full_sp,
                        "spacer_name": spacers[sp_idx][0][:100],
                        "strand": strand,
                        "mismatches": int(mm),
                        "spacer_length": len(full_sp),
                    })
        if i % 200 == 0 or i == len(obelisks):
            dt = time.time() - t0
            rate = i / max(1e-3, dt)
            eta = (len(obelisks) - i) / max(1e-3, rate)
            print(f"  {i}/{len(obelisks)} obelisks ({rate:.0f} seq/s, eta {eta:.0f}s) "
                  f"hits so far: {len(hits)}", flush=True)

    seen = set(); unique = []
    for h in hits:
        key = (h["obelisk_id"], h["obelisk_pos"], h["spacer_idx"], h["strand"])
        if key not in seen:
            seen.add(key); unique.append(h)

    obelisks_with_hit = set(h["obelisk_id"] for h in unique)

    out = {
        "tool": "crispr",
        "n_obelisks": len(obelisks),
        "n_spacers": len(spacers),
        "seed_k": int(seed_k),
        "max_mismatch": int(max_mismatch),
        "n_unique_hits": len(unique),
        "n_obelisks_with_hit": len(obelisks_with_hit),
        "pct_obelisks_with_hit": round(100 * len(obelisks_with_hit) / max(1, len(obelisks)), 4),
        "wall_time_s": round(time.time() - t0, 1),
        "hits": unique[:200],

    }

    JSON.mkdir(parents=True, exist_ok=True)
    with open(JSON / "crispr_spacer_hits.json", "w") as fh:
        json.dump(out, fh, indent=2)
    print(f"\nwrote {JSON / 'crispr_spacer_hits.json'}", flush=True)
    print(f"  unique hits: {len(unique)}", flush=True)
    print(f"  obelisks with at least one hit: {len(obelisks_with_hit)} "
          f"({100*len(obelisks_with_hit)/max(1,len(obelisks)):.2f}%)", flush=True)
    return out

if __name__ == "__main__":
    seed_k = int(sys.argv[1]) if len(sys.argv) > 1 else 20
    max_mm = int(sys.argv[2]) if len(sys.argv) > 2 else 2
    crispr(seed_k=seed_k, max_mismatch=max_mm)
evolve_HELP = "evolve: codon | phylo"

S_SANGUINIS = {
    "TTT": 17.7, "TTC": 18.0, "TTA": 13.7, "TTG":  9.4, "CTT": 14.0, "CTC": 13.7,
    "CTA":  4.0, "CTG": 16.6, "ATT": 26.4, "ATC": 22.0, "ATA":  6.0, "ATG": 25.7,
    "GTT": 18.5, "GTC": 11.9, "GTA":  9.6, "GTG": 18.6, "TCT": 11.5, "TCC":  9.5,
    "TCA": 10.4, "TCG":  6.4, "CCT":  9.7, "CCC":  6.9, "CCA":  9.3, "CCG":  9.0,
    "ACT": 11.5, "ACC": 13.3, "ACA": 11.4, "ACG":  6.9, "GCT": 16.0, "GCC": 14.7,
    "GCA": 17.3, "GCG":  6.9, "TAT": 18.7, "TAC": 12.6, "TAA":  1.5, "TAG":  0.4,
    "CAT": 12.6, "CAC": 11.0, "CAA": 22.2, "CAG": 13.9, "AAT": 24.4, "AAC": 21.3,
    "AAA": 51.8, "AAG": 14.6, "GAT": 36.7, "GAC": 19.4, "GAA": 39.2, "GAG": 18.0,
    "TGT":  3.3, "TGC":  4.9, "TGA":  1.4, "TGG": 10.1, "CGT": 11.5, "CGC":  8.7,
    "CGA":  3.7, "CGG":  4.1, "AGT": 13.0, "AGC": 12.4, "AGA":  6.8, "AGG":  3.5,
    "GGT": 23.3, "GGC": 17.4, "GGA": 22.2, "GGG":  8.7,
}

E_COLI = {
    "TTT": 22.3, "TTC": 16.0, "TTA": 13.9, "TTG": 13.7, "CTT": 11.9, "CTC": 11.1,
    "CTA":  3.9, "CTG": 52.8, "ATT": 30.5, "ATC": 25.1, "ATA":  4.4, "ATG": 27.5,
    "GTT": 18.4, "GTC": 15.3, "GTA": 10.9, "GTG": 26.4, "TCT":  8.5, "TCC":  8.6,
    "TCA":  7.8, "TCG":  8.9, "CCT":  7.0, "CCC":  5.5, "CCA":  8.4, "CCG": 23.2,
    "ACT":  8.9, "ACC": 23.4, "ACA":  7.0, "ACG": 14.3, "GCT": 15.5, "GCC": 25.5,
    "GCA": 20.2, "GCG": 33.0, "TAT": 16.2, "TAC": 12.2, "TAA":  2.0, "TAG":  0.3,
    "CAT": 12.9, "CAC":  9.7, "CAA": 15.3, "CAG": 28.8, "AAT": 17.7, "AAC": 21.5,
    "AAA": 33.2, "AAG": 10.5, "GAT": 32.2, "GAC": 19.0, "GAA": 39.1, "GAG": 17.7,
    "TGT":  5.2, "TGC":  6.5, "TGA":  1.0, "TGG": 13.9, "CGT": 20.9, "CGC": 21.6,
    "CGA":  3.6, "CGG":  5.4, "AGT":  8.8, "AGC": 16.0, "AGA":  2.1, "AGG":  1.2,
    "GGT": 25.5, "GGC": 27.1, "GGA":  8.0, "GGG": 11.3,
}

AA = {
    "TTT": "F", "TTC": "F", "TTA": "L", "TTG": "L", "CTT": "L", "CTC": "L",
    "CTA": "L", "CTG": "L", "ATT": "I", "ATC": "I", "ATA": "I", "ATG": "M",
    "GTT": "V", "GTC": "V", "GTA": "V", "GTG": "V", "TCT": "S", "TCC": "S",
    "TCA": "S", "TCG": "S", "AGT": "S", "AGC": "S", "CCT": "P", "CCC": "P",
    "CCA": "P", "CCG": "P", "ACT": "T", "ACC": "T", "ACA": "T", "ACG": "T",
    "GCT": "A", "GCC": "A", "GCA": "A", "GCG": "A", "TAT": "Y", "TAC": "Y",
    "CAT": "H", "CAC": "H", "CAA": "Q", "CAG": "Q", "AAT": "N", "AAC": "N",
    "AAA": "K", "AAG": "K", "GAT": "D", "GAC": "D", "GAA": "E", "GAG": "E",
    "TGT": "C", "TGC": "C", "TGG": "W", "CGT": "R", "CGC": "R", "CGA": "R",
    "CGG": "R", "AGA": "R", "AGG": "R", "GGT": "G", "GGC": "G", "GGA": "G",
    "GGG": "G", "TAA": "*", "TAG": "*", "TGA": "*",
}

def _norm(t):
    s = sum(t.values())
    if s == 0:
        return {c: 0.0 for c in t}
    return {c: v / s for c, v in t.items()}

def _euclid(p, q):
    return float(np.sqrt(sum((p.get(c, 0) - q.get(c, 0)) ** 2
                              for c in set(p) | set(q))))

def _cosine(p, q):
    keys = set(p) | set(q)
    pv = np.array([p.get(c, 0) for c in keys])
    qv = np.array([q.get(c, 0) for c in keys])
    return float(np.dot(pv, qv) / (np.linalg.norm(pv) * np.linalg.norm(qv)))

def _revc(s):
    return s.translate(str.maketrans("ACGT", "TGCA"))[::-1]

def _orf_codons(rna, prot):

    candidates = []
    for strand_idx, strand in enumerate((rna, _revc(rna))):
        for off in range(3):
            codons = [strand[i:i+3]
                      for i in range(off, len(strand) - 2, 3)]
            trans = "".join(AA.get(c, "X") for c in codons)
            j = trans.find(prot)
            if j < 0:
                continue
            cs = codons[j: j + len(prot)]
            candidates.append((cs[0] == "ATG", strand_idx, off, cs))
    if not candidates:
        return None

    candidates.sort(key=lambda t: (not t[0], t[1], t[2]))
    return candidates[0][3]

def codon():
    wb = openpyxl.load_workbook(RAW / "zheludev_mmc2.xlsx", read_only=True)
    ws = wb["Page 1"]
    obs = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or row[0] is None:
            if i > 0 and row[0] is None:
                break
            continue
        seq, prot = row[6], row[7]
        if not (isinstance(seq, str) and isinstance(prot, str)):
            continue
        if len(seq) < 200 or len(prot) < 50:
            continue
        obs.append({"id": str(row[0]), "seq": seq.upper().replace("U", "T"),
                    "orf": prot.upper()})
    print(f"loaded {len(obs):,} obelisks with RNA + ORF")

    print("extracting in-frame codons")
    cnt = Counter()
    n_ok = 0
    for ob in obs:
        cs = _orf_codons(ob["seq"], ob["orf"])
        if cs is not None:
            cnt.update(cs); n_ok += 1
    print(f"  {n_ok:,}/{len(obs):,} extracted")

    total = sum(cnt.values())
    per_1000 = {c: 1000 * v / total for c, v in cnt.items()}
    p_ob = _norm(per_1000)
    p_sa = _norm(S_SANGUINIS)
    p_ec = _norm(E_COLI)

    out = {
        "tool": "oblin1_codon_usage",
        "n_obelisks_total": len(obs),
        "n_ob_codons": n_ok,
        "total_codons": int(total),
        "codon_per_1000": {c: round(per_1000.get(c, 0), 2)
                                         for c in S_SANGUINIS},
        "S_sanguinis_SK36_per_1000": S_SANGUINIS,
        "E_coli_K12_per_1000": E_COLI,
        "distance_obelisk_vs_S_sanguinis_euclidean": _euclid(p_ob, p_sa),
        "distance_obelisk_vs_E_coli_euclidean": _euclid(p_ob, p_ec),
        "cosine_obelisk_vs_S_sanguinis": _cosine(p_ob, p_sa),
        "cosine_obelisk_vs_E_coli": _cosine(p_ob, p_ec),
    }
    p = JSON / "codon_usage.json"
    p.write_text(json.dumps(out, indent=2))
    print(f"\nwrote {p}")
    print(f"  Euclid (S. sanguinis) = "
          f"{out['distance_obelisk_vs_S_sanguinis_euclidean']:.4f}")
    print(f"  Euclid (E. coli)      = "
          f"{out['distance_obelisk_vs_E_coli_euclidean']:.4f}")
    print(f"  cosine (S. sanguinis) = "
          f"{out['cosine_obelisk_vs_S_sanguinis']:.4f}")
    print(f"  cosine (E. coli)      = "
          f"{out['cosine_obelisk_vs_E_coli']:.4f}")
    return out

def phylo():
    tree_path = JSON / "oblin1_tree.nwk"
    print(f"reading {tree_path.name}")
    tree = Phylo.read(tree_path, "newick")
    leaves = [t.name for t in tree.get_terminals()]
    print(f"  {len(leaves):,} leaves")

    fc = json.load(open(JSON / "catalog_features.json"))
    pass_by_id = {r["id"]: (r["frac_paired"] > 0.65
                            and r["sld_per100nt"] > 10
                            and r["max_loop"] <= 15)
                  for r in fc["OBELISK_FULL"]}

    pass_at_leaf = []
    matched = 0
    for L in leaves:
        if L in pass_by_id:
            pass_at_leaf.append(pass_by_id[L]); matched += 1
        else:
            pass_at_leaf.append(None)
    print(f"  {matched:,} leaves mapped to BRRC")

    matched_pass = [p for p in pass_at_leaf if p is not None]
    overall = 100 * sum(matched_pass) / len(matched_pass)
    print(f"  overall leaf BRRC pass: {overall:.1f}%")

    nodes = tree.get_nonterminals()
    print(f"\nper-clade pass for {len(nodes)} internal nodes")
    leaf_to_pass = dict(zip(leaves, pass_at_leaf))
    clades = []
    for node in nodes:
        st = [leaf_to_pass.get(t.name) for t in node.get_terminals()]
        st = [s for s in st if s is not None]
        if len(st) >= 5:
            clades.append({"n_leaves": len(st),
                           "pass_pct": 100 * sum(st) / len(st)})
    clades.sort(key=lambda d: -d["n_leaves"])

    sizes = [10, 50, 100, 500, 1000]
    distribution = {}
    for s in sizes:
        ge = [c["pass_pct"] for c in clades if c["n_leaves"] >= s]
        if ge:
            distribution[f"clades_with_>={s}_leaves"] = {
                "n_clades": len(ge),
                "median_pass_pct": float(np.median(ge)),
                "min_pass_pct": float(np.min(ge)),
                "max_pass_pct": float(np.max(ge)),
                "iqr_pass_pct": [float(np.percentile(ge, 25)),
                                  float(np.percentile(ge, 75))],
            }

    rng = np.random.default_rng(42)
    pass_arr = np.array([1 if s else 0 for s in matched_pass])
    matched_leaves = [L for L, s in zip(leaves, pass_at_leaf) if s is not None]
    obs_var = float(np.var([c["pass_pct"] for c in clades]))

    null_var = []
    for _ in range(200):
        perm = rng.permutation(pass_arr)
        L_to_p = {L: int(p) for L, p in zip(matched_leaves, perm)}
        v = []
        for node in nodes:
            ls = [L for L in (t.name for t in node.get_terminals())
                  if L in L_to_p]
            if len(ls) >= 5:
                v.append(100 * sum(L_to_p[L] for L in ls) / len(ls))
        null_var.append(float(np.var(v)))
    null_var = np.array(null_var)
    p_val = float((null_var >= obs_var).mean())

    out = {
        "tool": "phylo_oblin1",
        "n_leaves_in_tree": len(leaves),
        "n_leaves_mapped_to_BRRC": matched,
        "overall_BRRC_pass_pct_at_leaves": round(overall, 2),
        "n_internal_nodes_with_>=5_leaves": len(clades),
        "per_clade_distribution": distribution,
        "obs_clade_var": round(obs_var, 2),
        "perm_null_var_mean": round(float(null_var.mean()), 2),
        "perm_null_var_std": round(float(null_var.std()), 2),

        "perm_null_var_samples": [round(float(v), 4)
                                              for v in null_var],
        "permutation_p_value_obs_>=_null": p_val,
    }
    p = JSON / "phylo_oblin1.json"
    p.write_text(json.dumps(out, indent=2))
    print(f"\nwrote {p}")
    print(f"  observed clade variance: {obs_var:.1f}")
    print(f"  null mean variance:      {null_var.mean():.1f}")
    print(f"  p(obs >= null) = {p_val:.4f}")
    return out

evolve_CMDS = {"codon": codon, "phylo": phylo}

def evolve_main(argv=None):
    argv = sys.argv[1:] if argv is None else argv
    cmd = argv[0] if argv else "codon"
    if cmd not in evolve_CMDS:
        print(evolve_HELP, file=sys.stderr); sys.exit(2)
    evolve_CMDS[cmd]()

if __name__ == "__main__":
    evolve_main()
def _extval_topo(db: str):
    n = len(db)
    paired = sum(1 for c in db if c in "()")
    fp = paired / n if n else 0.0
    loops, cur = [], 0
    for c in db:
        if c == ".":
            cur += 1
        else:
            if cur > 0:
                loops.append(cur)
            cur = 0
    if cur > 0:
        loops.append(cur)
    sld = 100.0 * sum(1 for L in loops if L <= 5) / n if n else 0.0
    ml = max(loops) if loops else 0
    return round(fp, 4), round(sld, 2), int(ml)

def _extval_passes(fp: float, sld: float, ml: int) -> bool:
    return fp > 0.65 and sld > 10 and ml <= 15

def _extval_fasta(path: Path):
    out = []
    nm, buf = None, []
    with open(path) as fh:
        for line in fh:
            line = line.rstrip()
            if line.startswith(">"):
                if nm is not None:
                    out.append((nm, "".join(buf)))
                nm = line[1:]
                buf = []
            elif line:
                buf.append(line)
        if nm is not None:
            out.append((nm, "".join(buf)))
    return out

def _fold_circular(arg):
    nm, seq = arg
    su = seq.upper().replace("T", "U")
    su = "".join(c for c in su if c in "ACGUN")
    if not su or len(su) < 60 or len(su) > 5000:
        return {"id": nm.split()[0], "length": len(su), "error": "len_filter"}
    try:
        md = RNA.md()
        md.circ = 1
        fc = RNA.fold_compound(su, md)
        db, mfe = fc.mfe()
    except Exception as exc:
        return {"id": nm.split()[0], "length": len(su), "error": str(exc)}
    fp, sld, ml = _extval_topo(db)
    return {"id": nm.split()[0], "length": len(su),
            "mfe_per_nt": mfe / max(len(su), 1),
            "frac_paired": fp, "sld_per100nt": sld,
            "max_loop": ml, "passes_BRRC": _extval_passes(fp, sld, ml)}

ZIP_PATH = RAW / "hsob_supp_v2.zip"
HSOB_FASTA = RAW / "hsob_obelisks_new.fasta"

HSOB_CANDIDATES = [
    "Supplement_hsObl/sequence_files/obelisks_new.fasta",
    "obelisks_new.fasta",
    "sequence_files/obelisks_new.fasta",
    "Supplement_hsObl_v2.1/sequence_files/obelisks_new.fasta",
]
ALL_DEREP_CANDIDATES = [
    "Supplement_hsObl/cccRNA_cluster/allObelisks_derep_zheludev_tara_thisStudy.fasta",
    "cccRNA_cluster/allObelisks_derep_zheludev_tara_thisStudy.fasta",
    "Supplement_hsObl_v2.1/cccRNA_cluster/allObelisks_derep_zheludev_tara_thisStudy.fasta",
]

def _scan_zip_for_fasta(zip_path: Path) -> dict:
    out = {}
    with zipfile.ZipFile(zip_path) as z:
        for info in z.infolist():
            if info.is_dir():
                continue
            n = info.filename
            if n.endswith(".fasta") or n.endswith(".fa") or n.endswith(".fna"):
                out[n] = info.file_size
    return out

def _extract_member(zip_path: Path, member: str, dest: Path) -> bool:
    try:
        with zipfile.ZipFile(zip_path) as z:
            with z.open(member) as src, open(dest, "wb") as dst:
                dst.write(src.read())
        return True
    except (KeyError, zipfile.BadZipFile):
        return False

def locate_hsob_fasta() -> Path | None:
    if HSOB_FASTA.exists() and HSOB_FASTA.stat().st_size > 0:
        return HSOB_FASTA
    if not ZIP_PATH.exists():
        return None
    inv = _scan_zip_for_fasta(ZIP_PATH)

    candidates = (
        [m for m in HSOB_CANDIDATES if m in inv]
        + [m for m in inv if m.endswith("obelisks_new.fasta")]
        + [m for m in ALL_DEREP_CANDIDATES if m in inv]
        + [m for m in inv
           if "allObelisks_derep" in m and m.endswith(".fasta")]
    )
    if not candidates:
        print("[hsob] no obelisks_new.fasta in zip; available FASTAs:",
              flush=True)
        for k, v in sorted(inv.items()):
            print(f"   {k} ({v} B)")
        return None
    pick = candidates[0]
    print(f"[hsob] extracting {pick}", flush=True)
    if _extract_member(ZIP_PATH, pick, HSOB_FASTA):
        return HSOB_FASTA
    return None

def primary() -> dict:
    fa = locate_hsob_fasta()
    if fa is None:
        print("[primary] HsOb FASTA not located; aborting primary mode",
              flush=True)
        return {}
    seqs = _extval_fasta(fa)

    seen = set()
    keep = []
    for nm, s in seqs:
        su = s.upper().replace("T", "U")
        if not (500 <= len(su) <= 2500):
            continue
        if su in seen:
            continue
        seen.add(su)
        keep.append((nm, su))
    print(f"[primary] loaded {len(seqs):,} HsOb seqs; "
          f"{len(keep):,} after dedup + length filter", flush=True)

    workers = max(1, mp.cpu_count() - 1)
    t0 = time.time()
    rows = []
    with mp.Pool(workers) as pool:
        for i, r in enumerate(pool.imap_unordered(_fold_circular, keep,
                                                  chunksize=4), 1):
            rows.append(r)
            if i % 250 == 0 or i == len(keep):
                rate = i / max(time.time() - t0, 1e-6)
                print(f"  folded {i:,}/{len(keep):,} ({rate:.1f}/s)",
                      flush=True)

    ok = [r for r in rows if "error" not in r]
    npass = sum(1 for r in ok if r["passes_BRRC"])
    n_fp = sum(1 for r in ok if r["frac_paired"] > 0.65)
    n_sld = sum(1 for r in ok if r["sld_per100nt"] > 10)
    n_ml = sum(1 for r in ok if r["max_loop"] <= 15)

    summary = {
        "mode": "primary_external_HsOb",
        "source": {
            "repository": "Zenodo",
            "doi": "10.5281/zenodo.18551497",
            "record_url": "https://zenodo.org/records/18551497",
            "archive": "Supplement_hsObl_v2.1.zip",
            "fasta_member_used": str(fa.name),
            "paper": "Pichler/Urayama et al., Nat Commun 17 (2026)",
        },
        "n_input_sequences": len(seqs),
        "n_after_dedup_len": len(keep),
        "n_folded_ok": len(ok),
        "length_window_nt": [500, 2500],
        "BRRC_full_pass": {
            "n": npass, "pct": round(100 * npass / max(len(ok), 1), 2)},
        "by_criterion_pass": {
            "fp_gt065_pct":
                round(100 * n_fp / max(len(ok), 1), 2),
            "sld_gt10_pct":
                round(100 * n_sld / max(len(ok), 1), 2),
            "max_loop_le_15_pct":
                round(100 * n_ml / max(len(ok), 1), 2),
        },
        "comparison_to_Zheludev_linear_BRRC_pct": 87.7,
        "means": {
            "frac_paired": float(np.mean([r["frac_paired"] for r in ok]))
                if ok else None,
            "sld_per100nt":
                float(np.mean([r["sld_per100nt"]
                               for r in ok])) if ok else None,
            "max_loop": float(np.mean([r["max_loop"] for r in ok]))
                if ok else None,
            "mfe_per_nt": float(np.mean([r["mfe_per_nt"] for r in ok]))
                if ok else None,
        },
    }
    return {"summary": summary, "rows": rows}

def _kmers(seq: str, k: int = 4) -> set:
    s = seq.upper().replace("T", "U")
    return {s[i:i+k] for i in range(len(s) - k + 1) if "N" not in s[i:i+k]}

def _jaccard(a: set, b: set) -> float:
    if not a or not b:
        return 0.0
    inter = len(a & b)
    return inter / (len(a) + len(b) - inter)

def _cluster_single_linkage(ids, kmer_sets, threshold: float = 0.30):
    n = len(ids)
    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    sizes = [len(s) for s in kmer_sets]
    order = sorted(range(n), key=lambda i: sizes[i])

    win = 64
    rng = np.random.default_rng(42)
    for idx, i in enumerate(order):

        for j in order[idx + 1: idx + 1 + win]:
            if _jaccard(kmer_sets[i], kmer_sets[j]) >= threshold:
                union(i, j)

        for j in rng.choice(n, size=min(8, n), replace=False):
            if j == i:
                continue
            if _jaccard(kmer_sets[i], kmer_sets[j]) >= threshold:
                union(i, j)

    root = [find(i) for i in range(n)]
    clusters = defaultdict(list)
    for i, r in enumerate(root):
        clusters[r].append(i)
    return list(clusters.values())

def fallback() -> dict:
    seqs = _extval_fasta(PROC / "obelisks_zheludev_catalog.fasta")
    ids = [nm.split()[0] for nm, _ in seqs]
    raws = [s for _, s in seqs]
    print(f"[fallback] loaded {len(seqs):,} Zheludev obelisks", flush=True)

    print("[fallback] computing 4-mer sets ...", flush=True)
    ksets = [_kmers(s, 4) for s in raws]
    print("[fallback] clustering at Jaccard >= 0.30 ...", flush=True)
    clusters = _cluster_single_linkage(ids, ksets, threshold=0.30)
    sizes = sorted([len(c) for c in clusters], reverse=True)
    print(f"  {len(clusters):,} clusters; sizes top10: {sizes[:10]}",
          flush=True)

    rng = np.random.default_rng(42)
    cl = list(clusters)
    rng.shuffle(cl)
    cl.sort(key=lambda c: -len(c))
    bin_a, bin_b = [], []
    na = nb = 0
    for c in cl:
        if na <= nb:
            bin_a.append(c)
            na += len(c)
        else:
            bin_b.append(c)
            nb += len(c)
    train_idx = sum(bin_a, [])
    test_idx = sum(bin_b, [])
    print(f"  train={len(train_idx):,}  test={len(test_idx):,} "
          "(cluster-disjoint)", flush=True)

    fc_path = JSON / "catalog_features.json"
    fc_rows = []
    if fc_path.exists():
        fc = json.loads(fc_path.read_text())
        rows = fc.get("OBELISK_FULL") or fc.get("obelisk_full") or []

        by_id = {r["id"] if "id" in r else r.get("name"): r for r in rows}

        if not by_id or list(by_id.keys())[0] is None:
            by_id = {ids[i]: rows[i] for i in range(min(len(ids), len(rows)))}
        for i in train_idx:
            fc_rows.append(("train", ids[i], by_id.get(ids[i])))
        for i in test_idx:
            fc_rows.append(("test", ids[i], by_id.get(ids[i])))

    if not fc_rows or any(r[2] is None for r in fc_rows):

        print("[fallback] catalog_features.json missing or mis-keyed; "
              "folding on the fly", flush=True)
        workers = max(1, mp.cpu_count() - 1)
        args = [(ids[i], raws[i]) for i in train_idx + test_idx]
        labels = (["train"] * len(train_idx)) + (["test"] * len(test_idx))
        with mp.Pool(workers) as pool:
            folded = list(pool.imap(_fold_circular, args, chunksize=8))
        fc_rows = list(zip(labels, [ids[i] for i in train_idx + test_idx],
                           folded))

    train_rows = [r for lab, _, r in fc_rows if lab == "train" and r]
    test_rows = [r for lab, _, r in fc_rows if lab == "test" and r]

    def _pass_stats(rs):
        n = len(rs)
        if not n:
            return {"n": 0, "BRRC_pass_pct": None}

        def _is_pass(r):

            if "passes_BRRC" in r:
                return bool(r["passes_BRRC"])
            return _extval_passes(r["frac_paired"],
                           r.get("sld_per100nt",
                                 r.get("sld", 0)),
                           r["max_loop"])
        npass = sum(1 for r in rs if _is_pass(r))
        return {"n": n, "BRRC_pass_pct": round(100 * npass / n, 2),
                "frac_paired_mean":
                    float(np.mean([r["frac_paired"] for r in rs])),
                "max_loop_mean":
                    float(np.mean([r["max_loop"] for r in rs])),
                "sld_per100nt_mean":
                    float(np.mean(
                        [r.get("sld_per100nt", r.get("sld", 0))
                         for r in rs]))}

    train_stats = _pass_stats(train_rows)
    test_stats = _pass_stats(test_rows)
    summary = {
        "mode": "fallback_cross_cluster_holdout",
        "input_set": "Zheludev 2024 catalog (n=5169)",
        "method": "4-mer Jaccard clustering, 50/50 cluster-disjoint split",
        "n_clusters": len(clusters),
        "largest_cluster_size": sizes[0] if sizes else 0,
        "singletons": int(sum(1 for s in sizes if s == 1)),
        "train": train_stats,
        "test": test_stats,
        "comparison_to_full_catalog_linear_BRRC_pct": 87.7,
    }
    return {"summary": summary}

def write(payload: dict) -> Path:
    out = JSON / "extval.json"
    out.write_text(json.dumps(payload, indent=2))
    print(f"\nwrote {out}", flush=True)
    return out

def external_validation_main(argv=None):
    argv = sys.argv[1:] if argv is None else argv
    mode = argv[0] if argv else "auto"

    if mode == "primary":
        payload = primary()
        if not payload:
            payload = {"summary": {"mode": "primary_FAILED", "reason": "HsOb_FASTA_missing"}}
        write(payload)
        return

    if mode == "fallback":
        payload = fallback()
        write(payload)
        return

    fa = locate_hsob_fasta()
    if fa is not None and fa.stat().st_size > 100_000:
        payload = primary()
        if payload:
            write(payload)
            return
    print("[auto] HsOb FASTA unavailable; running fallback", flush=True)
    write(fallback())

if __name__ == "__main__":
    external_validation_main()
def _loop_motifs_circ_fold(seq):
    s = seq.upper().replace("T", "U")
    md = RNA.md(); md.circ = 1
    fc = RNA.fold_compound(s, md)
    db, _ = fc.mfe()
    return db

def _context_kmers(seq, db, k=6, context="loop"):
    s = seq.upper().replace("U", "T")
    n = len(s)
    if n != len(db) or n < k:
        return []
    out = []
    for i in range(n):

        start = (i - k // 2) % n
        kmer_seq = "".join(s[(start + j) % n] for j in range(k))
        if "N" in kmer_seq or set(kmer_seq) - set("ACGT"):
            continue
        if context == "loop" and db[i] == ".":
            out.append(kmer_seq)
        elif context == "stem" and db[i] in "()":
            out.append(kmer_seq)
        elif context == "all":
            out.append(kmer_seq)
    return out

def _bh_fdr(pvals):
    n = len(pvals)
    if n == 0:
        return []
    order = sorted(range(n), key=lambda i: pvals[i])
    ranks = [0] * n
    for r, i in enumerate(order):
        ranks[i] = r + 1
    q = [0.0] * n
    prev = 1.0
    for i in reversed(order):
        adj = pvals[i] * n / ranks[i]
        prev = min(prev, adj)
        q[i] = prev
    return q

def _chi2(a, b, c, d):
    n = a + b + c + d
    if n == 0:
        return 0.0, 1.0
    e_a = (a + b) * (a + c) / n
    e_b = (a + b) * (b + d) / n
    e_c = (c + d) * (a + c) / n
    e_d = (c + d) * (b + d) / n
    chi2 = 0.0
    for obs, exp in [(a, e_a), (b, e_b), (c, e_c), (d, e_d)]:
        if exp > 1e-9:
            chi2 += (obs - exp) ** 2 / exp

    from scipy.stats import chi2 as scchi2
    p = 1.0 - scchi2.cdf(chi2, df=1)
    return chi2, p

def motifs(n_obelisk=300, k=6, context="loop", seed=42):
    rng = random.Random(seed)

    print("loading sequences", flush=True)
    ob_all = _fasta(PROC / "obelisks_zheludev_catalog.fasta")
    ob_all = [(nm, s) for nm, s in ob_all if 700 <= len(s) <= 1700]
    rng.shuffle(ob_all)
    ob_sample = ob_all[:n_obelisk]

    hdv = []
    for p in (PROC / "delta_genomes.fasta", PROC / "delta_full_genomes.fasta"):
        if p.exists():
            hdv = _fasta(p)
            break
    print(f"  obelisks: {len(ob_sample)}, HDV: {len(hdv)}, viroids: {len(VIROIDS)}",
          flush=True)

    print("folding and counting context k-mers", flush=True)
    ob_kmers = Counter()
    ob_totals = 0
    for i, (nm, s) in enumerate(ob_sample, 1):
        try:
            db = _loop_motifs_circ_fold(s)
            kms = _context_kmers(s, db, k=k, context=context)
            ob_kmers.update(kms); ob_totals += len(kms)
        except Exception:
            continue
        if i % 50 == 0:
            print(f"  obelisk {i}/{len(ob_sample)}", flush=True)

    other_kmers = Counter()
    other_totals = 0
    for nm, s in hdv:
        try:
            db = _loop_motifs_circ_fold(s)
            kms = _context_kmers(s, db, k=k, context=context)
            other_kmers.update(kms); other_totals += len(kms)
        except Exception:
            continue
    for nm, s in VIROIDS.items():
        try:
            db = _loop_motifs_circ_fold(s)
            kms = _context_kmers(s, db, k=k, context=context)
            other_kmers.update(kms); other_totals += len(kms)
        except Exception:
            continue

    print(f"  obelisk loop k-mers: {ob_totals}, other: {other_totals}", flush=True)

    rows = []
    all_kmers = set(ob_kmers) | set(other_kmers)
    for km in all_kmers:
        a = ob_kmers.get(km, 0)
        c = other_kmers.get(km, 0)
        b = ob_totals - a
        d = other_totals - c
        if a < 20 and c < 20:
            continue
        chi2, p = _chi2(a, b, c, d)
        obelisk_freq = a / max(1, ob_totals)
        other_freq = c / max(1, other_totals)
        log_or = math.log2((obelisk_freq + 1e-9) / (other_freq + 1e-9))
        rows.append({
            "kmer": km,
            "obelisk_count": a,
            "ob_freq_per_M": round(1e6 * obelisk_freq, 2),
            "other_count": c,
            "other_freq_per_million": round(1e6 * other_freq, 2),
            "log2_OR": round(log_or, 3),
            "chi2": round(chi2, 2),
            "p_value": p,
        })

    if rows:
        qvals = _bh_fdr([r["p_value"] for r in rows])
        for r, q in zip(rows, qvals):
            r["q_value"] = q
            r["p_value"] = float(r["p_value"])
        rows.sort(key=lambda r: (-abs(r["log2_OR"]), r["q_value"]))

    for r in rows:
        r["p_value"] = float(f"{r['p_value']:.3e}")
        r["q_value"] = float(f"{r['q_value']:.3e}")

    top_enriched = [r for r in rows if r["log2_OR"] > 0 and r["q_value"] < 0.01][:20]
    top_depleted = [r for r in rows if r["log2_OR"] < 0 and r["q_value"] < 0.01][:20]

    out = {
        "tool": "motifs",
        "k": k,
        "context": context,
        "n_obelisks": len(ob_sample),
        "n_other": len(hdv) + len(VIROIDS),
        "n_obelisk_kmers_total": ob_totals,
        "n_other_kmers_total": other_totals,
        "n_kmers_tested": len(rows),
        "n_sig_enr_fdr01": sum(
            1 for r in rows if r["log2_OR"] > 0 and r["q_value"] < 0.01),
        "n_sig_depl_fdr01": sum(
            1 for r in rows if r["log2_OR"] < 0 and r["q_value"] < 0.01),
        "top20_enriched_loops": top_enriched,
        "top20_depleted_loops": top_depleted,

    }

    JSON.mkdir(parents=True, exist_ok=True)
    with open(JSON / "loop_motif_enrichment.json", "w") as fh:
        json.dump(out, fh, indent=2)
    print(f"\nwrote {JSON / 'loop_motif_enrichment.json'}", flush=True)
    print(f"\ntop 10 enriched in obelisk {context} ({k}-mers):", flush=True)
    for r in top_enriched[:10]:
        print(f"  {r['kmer']}  log2OR={r['log2_OR']:+.2f}  q={r['q_value']:.2e}  "
              f"obelisk={r['ob_freq_per_M']:.0f}/M  "
              f"other={r['other_freq_per_million']:.0f}/M", flush=True)
    return out

if __name__ == "__main__":
    n = int(sys.argv[1]) if len(sys.argv) > 1 else 300
    k = int(sys.argv[2]) if len(sys.argv) > 2 else 6
    ctx = sys.argv[3] if len(sys.argv) > 3 else "loop"
    motifs(n_obelisk=n, k=k, context=ctx)
SYN_TABLE = {
    "TTT": "TTC", "TTC": "TTT", "TTA": "TTG", "TTG": "TTA",
    "CTT": "CTC", "CTC": "CTA", "CTA": "CTG", "CTG": "CTT",
    "ATT": "ATC", "ATC": "ATA", "ATA": "ATT",
    "GTT": "GTC", "GTC": "GTA", "GTA": "GTG", "GTG": "GTT",
    "TCT": "TCC", "TCC": "TCA", "TCA": "TCG", "TCG": "TCT",
    "CCT": "CCC", "CCC": "CCA", "CCA": "CCG", "CCG": "CCT",
    "ACT": "ACC", "ACC": "ACA", "ACA": "ACG", "ACG": "ACT",
    "GCT": "GCC", "GCC": "GCA", "GCA": "GCG", "GCG": "GCT",
    "TAT": "TAC", "TAC": "TAT",
    "CAT": "CAC", "CAC": "CAT", "CAA": "CAG", "CAG": "CAA",
    "AAT": "AAC", "AAC": "AAT", "AAA": "AAG", "AAG": "AAA",
    "GAT": "GAC", "GAC": "GAT", "GAA": "GAG", "GAG": "GAA",
    "TGT": "TGC", "TGC": "TGT",
    "CGT": "CGC", "CGC": "CGA", "CGA": "CGG", "CGG": "CGT",
    "AGT": "AGC", "AGC": "AGT", "AGA": "AGG", "AGG": "AGA",
    "GGT": "GGC", "GGC": "GGA", "GGA": "GGG", "GGG": "GGT",
}

def _fold_circ(seq):
    s = seq.upper().replace("T", "U")
    md = RNA.md(); md.circ = 1
    fc = RNA.fold_compound(s, md)
    db, _ = fc.mfe()
    return db

def _find_orf(s):
    s = s.upper().replace("U", "T")
    stops = {"TAA", "TAG", "TGA"}
    best = None
    for off in range(3):
        i = off
        while i + 3 <= len(s):
            if s[i:i + 3] == "ATG":
                j = i
                while j + 3 <= len(s):
                    if s[j:j + 3] in stops:
                        if best is None or (j + 3 - i) > (best[1] - best[0]):
                            best = (i, j + 3)
                        break
                    j += 3
                i = j + 3
            else:
                i += 3
    return best

def rna_mut(n_obelisks=10, n_random_paired=12, n_random_silent=12, seed=42):
    rng = random.Random(seed)
    print("loading obelisks", flush=True)
    seqs = _fasta(PROC / "obelisks_zheludev_catalog.fasta")
    keep = [(nm, s) for nm, s in seqs if 700 <= len(s) <= 1700]
    rng.shuffle(keep)
    sample = keep[:n_obelisks]

    results = []
    t0 = time.time()
    for i, (nm, s) in enumerate(sample, 1):
        s = s.upper().replace("U", "T")
        try:
            db0 = _fold_circ(s)
            fp0, sld0, ml0 = _topo(db0)
            pass0 = _passes(fp0, sld0, ml0)
        except Exception:
            continue

        orf = _find_orf(s)
        if orf is None:
            continue
        orf_a, orf_b = orf

        paired_in_orf = [k for k in range(orf_a, orf_b)
                         if k < len(db0) and db0[k] in "()"]
        rng.shuffle(paired_in_orf)
        paired_in_orf = paired_in_orf[:n_random_paired]

        codon_starts = [orf_a + 3 * k for k in range((orf_b - orf_a) // 3)
                        if orf_a + 3 * (k + 1) <= orf_b]
        synonymizable = [k for k in codon_starts
                         if s[k:k + 3] in SYN_TABLE]
        rng.shuffle(synonymizable)
        synonymizable = synonymizable[:n_random_silent]

        paired_results = []
        for pos in paired_in_orf:
            orig = s[pos]
            new = {"A": "C", "C": "A", "G": "T", "T": "A"}.get(orig, "A")
            mut = s[:pos] + new + s[pos + 1:]
            try:
                db = _fold_circ(mut)
                fp, sld, ml = _topo(db)
                p = _passes(fp, sld, ml)
            except Exception:
                continue
            paired_results.append({
                "pos": pos, "orig": orig, "new": new,
                "delta_frac_paired": round(fp - fp0, 4),
                "delta_sld": round(sld - sld0, 2),
                "delta_max_loop": int(ml - ml0),
                "brrc_pass": bool(p),
            })

        silent_results = []
        for cstart in synonymizable:
            orig_codon = s[cstart:cstart + 3]
            new_codon = SYN_TABLE.get(orig_codon, orig_codon)
            if new_codon == orig_codon:
                continue
            mut = s[:cstart] + new_codon + s[cstart + 3:]
            try:
                db = _fold_circ(mut)
                fp, sld, ml = _topo(db)
                p = _passes(fp, sld, ml)
            except Exception:
                continue
            silent_results.append({
                "codon_start": cstart, "orig": orig_codon, "new": new_codon,
                "delta_frac_paired": round(fp - fp0, 4),
                "delta_sld": round(sld - sld0, 2),
                "delta_max_loop": int(ml - ml0),
                "brrc_pass": bool(p),
            })

        results.append({
            "id": nm,
            "wild_type": {
                "frac_paired": fp0, "sld": sld0, "max_loop": ml0, "brrc_pass": bool(pass0),
            },
            "orf_span": [orf_a, orf_b],
            "paired_disruption": paired_results,
            "silent_coding": silent_results,
        })

        if i % 2 == 0 or i == len(sample):
            print(f"  {i}/{len(sample)} obelisks ({(time.time() - t0):.0f}s)", flush=True)

    paired_pass = sum(1 for r in results for x in r["paired_disruption"] if x["brrc_pass"])
    paired_total = sum(len(r["paired_disruption"]) for r in results)
    silent_pass = sum(1 for r in results for x in r["silent_coding"] if x["brrc_pass"])
    silent_total = sum(len(r["silent_coding"]) for r in results)

    summary = {
        "n_obelisks": len(results),
        "paired_disruption": {
            "n": paired_total,
            "n_pass": paired_pass,
            "pass_pct": round(100 * paired_pass / max(1, paired_total), 2),
        },
        "silent_coding": {
            "n": silent_total,
            "n_pass": silent_pass,
            "pass_pct": round(100 * silent_pass / max(1, silent_total), 2),
        },

    }

    out = {
        "tool": "rna_mut",
        "seed": seed,
        "summary": summary,
        "per_obelisk_results": results,
    }
    JSON.mkdir(parents=True, exist_ok=True)
    with open(JSON / "rna_mutation_scan.json", "w") as fh:
        json.dump(out, fh, indent=2)
    print(f"\nwrote {JSON / 'rna_mutation_scan.json'}", flush=True)
    print(f"summary: {json.dumps(summary, indent=2)}", flush=True)
    return out

if __name__ == "__main__":
    n = int(sys.argv[1]) if len(sys.argv) > 1 else 10
    rna_mut(n_obelisks=n)
def _zscore(X):
    X = np.asarray(X, dtype=float)
    mu = X.mean(axis=0)
    sd = X.std(axis=0)
    sd[sd < 1e-9] = 1.0
    return (X - mu) / sd

def _pdist(X):
    X = np.asarray(X, dtype=float)
    n = X.shape[0]
    out = []
    for i in range(n - 1):
        for j in range(i + 1, n):
            out.append(float(np.linalg.norm(X[i] - X[j])))
    return np.asarray(out)

def _rank(x):
    x = np.asarray(x, dtype=float)
    order = np.argsort(x)
    ranks = np.empty_like(order, dtype=float)
    ranks[order] = np.arange(len(x))

    return ranks + 1.0

def _spearman(a, b):
    ra = _rank(a); rb = _rank(b)
    if ra.std() < 1e-9 or rb.std() < 1e-9:
        return 0.0
    return float(np.corrcoef(ra, rb)[0, 1])

def _mantel(D_rna, D_prot, n_perm=999, seed=42):
    rng = random.Random(seed)
    n = int((1 + math.sqrt(1 + 8 * len(D_rna))) / 2)

    def to_sq(v):
        M = np.zeros((n, n))
        idx = 0
        for i in range(n - 1):
            for j in range(i + 1, n):
                M[i, j] = M[j, i] = v[idx]; idx += 1
        return M

    B = to_sq(D_prot)
    rho_obs = _spearman(D_rna, D_prot)

    null_rhos = []
    indices = list(range(n))
    for _ in range(n_perm):
        rng.shuffle(indices)
        B_perm = B[np.ix_(indices, indices)]

        v = []
        for i in range(n - 1):
            for j in range(i + 1, n):
                v.append(B_perm[i, j])
        null_rhos.append(_spearman(D_rna, np.asarray(v)))

    null_rhos = np.asarray(null_rhos)
    p_two = (np.sum(np.abs(null_rhos) >= abs(rho_obs)) + 1) / (n_perm + 1)
    return rho_obs, float(p_two), null_rhos

def coupling(n_sample=300, seed=42, n_perm=999):
    rng = random.Random(seed)

    print("loading RNA features", flush=True)
    fc = json.load(open(JSON / "catalog_features.json"))
    rna_rows = {r["id"]: r for r in fc["OBELISK_FULL"]}

    print("loading Oblin-1 protein features", flush=True)
    prot_rows = {r["id"]: r for r in json.load(open(JSON / "oblin1_conservation.json"))}

    common = sorted(set(rna_rows) & set(prot_rows))
    print(f"  {len(common)} obelisks have both RNA and Oblin-1 features", flush=True)
    if len(common) < 20:
        print("not enough paired data", flush=True)
        return None

    rng.shuffle(common)
    sample_ids = common[:n_sample]
    print(f"  sampling n={len(sample_ids)} for Mantel test", flush=True)

    rna_features = ["frac_paired", "sld_per100nt", "max_loop",
                    "mfe_per_nt", "L"]
    prot_features = ["pos_frac", "neg_frac", "net_charge", "aromatic",
                     "disorder", "hydrophobicity", "max_basic_window10"]

    X_rna = np.array([
        [float(rna_rows[i].get(k, 0.0)) for k in rna_features]
        for i in sample_ids
    ])
    X_prot = np.array([
        [float(prot_rows[i].get(k, 0.0)) for k in prot_features]
        for i in sample_ids
    ])

    Z_rna = _zscore(X_rna)
    Z_prot = _zscore(X_prot)

    D_rna = _pdist(Z_rna)
    D_prot = _pdist(Z_prot)

    print(f"  computing Mantel correlation over {len(D_rna)} pairs and "
          f"{n_perm} permutations", flush=True)
    t0 = time.time()
    rho, p, null_rhos = _mantel(D_rna, D_prot, n_perm=n_perm, seed=seed)
    dt = time.time() - t0
    print(f"  Mantel rho = {rho:.4f}, p = {p:.4f} (wall {dt:.1f}s)", flush=True)

    out = {
        "tool": "coupling",
        "n_obelisks_paired": len(sample_ids),
        "n_pairs": int(len(D_rna)),
        "n_permutations": int(n_perm),
        "seed": int(seed),
        "rna_features": rna_features,
        "protein_features": prot_features,
        "mantel_spearman_rho": round(float(rho), 4),
        "mantel_p_two_sided": round(float(p), 4),
        "null_rho_mean": round(float(null_rhos.mean()), 4),
        "null_rho_sd": round(float(null_rhos.std()), 4),
        "null_rho_p95": round(float(np.percentile(null_rhos, 95)), 4),
        "null_rho_max": round(float(null_rhos.max()), 4),
        "wall_time_s": round(dt, 1),

    }
    JSON.mkdir(parents=True, exist_ok=True)
    with open(JSON / "rna_oblin_coupling.json", "w") as fh:
        json.dump(out, fh, indent=2)
    print(f"wrote {JSON / 'rna_oblin_coupling.json'}", flush=True)
    print(json.dumps(out, indent=2), flush=True)
    return out

if __name__ == "__main__":
    n = int(sys.argv[1]) if len(sys.argv) > 1 else 300
    coupling(n_sample=n)
rnap_HELP = "rnap: promoter | sixs"

M35 = "TTGACA"
M10 = "TATAAT"
SP_LO, SP_HI = 15, 19

def _hamming(a, b):
    return sum(1 for x, y in zip(a, b) if x != y)

def _revcomp(s):
    return s.translate(str.maketrans("ACGTN", "TGCAN"))[::-1]

def _scan(seq, mm35=1, mm10=1):
    seq = seq.upper().replace("U", "T")
    n_pairs, best = 0, 0
    for s in (seq, _revcomp(seq)):
        L = len(s)
        ten = [(i, _hamming(s[i:i+len(M10)], M10))
               for i in range(L - len(M10) + 1)]
        ten = [(i, mm) for i, mm in ten if mm <= mm10]
        for i, mm10v in ten:
            lo = i - SP_HI - len(M35)
            hi = i - SP_LO - len(M35)
            for j in range(max(0, lo), hi + 1):
                mm35v = _hamming(s[j:j+len(M35)], M35)
                if mm35v <= mm35:
                    n_pairs += 1
                    score = (len(M35) + len(M10)) - (mm35v + mm10v)
                    if score > best:
                        best = score
    return n_pairs, best

def _rnap_fasta(path):
    out = []
    nm, buf = None, []
    with open(path) as fh:
        for line in fh:
            if line.startswith(">"):
                if nm is not None:
                    out.append("".join(buf))
                nm = line[1:].strip(); buf = []
            else:
                buf.append(line.strip())
        if nm is not None:
            out.append("".join(buf))
    return out

def _shuffle_di(seq, rng):

    seq = seq.upper().replace("U", "T")
    if len(seq) < 4:
        return seq
    edges = {b: [] for b in set(seq)}
    for a, b in zip(seq, seq[1:]):
        edges.setdefault(a, []).append(b)
    last = seq[-1]
    for _ in range(8):
        last_edge = {}
        nodes = [n for n in edges if n != last and edges[n]]
        for n in nodes:
            last_edge[n] = rng.choice(edges[n])
        rem = {n: list(edges[n]) for n in edges}
        for n in nodes:
            rem[n].remove(last_edge[n])
            rng.shuffle(rem[n])
            rem[n].append(last_edge[n])
        out, cur, ok = [seq[0]], seq[0], True
        for _ in range(len(seq) - 1):
            if not rem.get(cur):
                ok = False; break
            nxt = rem[cur].pop(0)
            out.append(nxt); cur = nxt
        if ok and len(out) == len(seq):
            return "".join(out)
    chars = list(seq); rng.shuffle(chars); return "".join(chars)

def promoter():
    seqs = _rnap_fasta(PROC / "obelisks_zheludev_catalog.fasta")
    print(f"  {len(seqs):,} obelisks")
    rng = np.random.default_rng(42)

    print("scanning obelisks for sigma-70 -35/-10 pairs (<=1 mm each)")
    ob_n, ob_b = [], []
    for s in seqs:
        n, b = _scan(s)
        ob_n.append(n); ob_b.append(b)

    print("dinucleotide-shuffled controls (1 per obelisk, seed 42)")
    sh_n, sh_b = [], []
    for i, s in enumerate(seqs):
        n, b = _scan(_shuffle_di(s, rng))
        sh_n.append(n); sh_b.append(b)
        if (i + 1) % 1000 == 0:
            print(f"  {i+1:,}/{len(seqs):,}")

    ob_n, ob_b = np.array(ob_n), np.array(ob_b)
    sh_n, sh_b = np.array(sh_n), np.array(sh_b)

    summary = {
        "tool": "sigma70_scan",
        "n_obelisks": int(len(ob_n)),
        "n_shuffles": int(len(sh_n)),
        "obelisk_pct_with_pair": float(100 * (ob_n >= 1).mean()),
        "shuffle_pct_with_pair": float(100 * (sh_n >= 1).mean()),
        "ob_mean_pairs": float(ob_n.mean()),
        "shuf_mean_pairs": float(sh_n.mean()),
        "ob_pct_perfect": float(100 * (ob_b == 12).mean()),
        "shuf_pct_perfect": float(100 * (sh_b == 12).mean()),
    }
    summary["enrichment_pct_with_pair"] = float(
        summary["obelisk_pct_with_pair"]
        / max(summary["shuffle_pct_with_pair"], 1e-9))

    bins = np.arange(0, 9)
    summary["histogram_obelisk_pair_counts"] = {
        str(int(b)): int(((ob_n == b) if b < 8 else (ob_n >= 8)).sum())
        for b in bins
    }
    summary["histogram_shuffle_pair_counts"] = {
        str(int(b)): int(((sh_n == b) if b < 8 else (sh_n >= 8)).sum())
        for b in bins
    }

    out = JSON / "promoter_scan.json"
    out.write_text(json.dumps(summary, indent=2))
    print(f"\nwrote {out}")
    print(f"  obelisks with >=1 pair: {summary['obelisk_pct_with_pair']:.1f}%")
    print(f"  shuffles with >=1 pair: {summary['shuffle_pct_with_pair']:.1f}%")
    print(f"  enrichment:             {summary['enrichment_pct_with_pair']:.2f}x")
    return summary

SIXS_ECOLI = (
    "ATTTCTCTGAGATGTTCGCAAGCGGGCCAGTCCCCTGAGCCCGAAACCCC"
    "TTAACGTAGTGAGTAGAACTCCGGCATGAATGCATCAACGTCAAAACTCT"
    "CCCATCAAATAACTAGCAATCATCAACGACAGAATAGCAAATCAGTGAGC"
    "AGCATGCATCATGCAACGTAATCATATCAGGTACAATTTGGTAACGT"
)

def _rnap_topo(db):

    paired = sum(1 for c in db if c in "()")
    fp = paired / len(db)
    loops, cur = [], 0
    for c in db:
        if c == ".":
            cur += 1
        else:
            if cur > 0:
                loops.append(cur); cur = 0
    if cur > 0:
        loops.append(cur)
    sld = 100 * sum(1 for L in loops if L <= 5) / len(db)
    ml = max(loops) if loops else 0
    return fp, sld, ml

def sixs():
    db, mfe = RNA.fold(SIXS_ECOLI)
    fp, sld, ml = _rnap_topo(db)
    print(f"E. coli 6S RNA (ssrS, {len(SIXS_ECOLI)} nt)")
    print(f"  MFE = {mfe:.1f} kcal/mol")
    print(f"  fp={fp:.3f}  sld={sld:.2f}  max_loop={ml}")

    passes = {
        "frac_paired_>0.65": fp > 0.65,
        "small_loop_density_>10": sld > 10,
        "max_loop_<=15": ml <= 15,
    }
    passes["all_three"] = all(passes.values())

    fc = json.load(open(JSON / "catalog_features.json"))
    ob = fc["OBELISK_FULL"]
    ob_fp = np.array([r["frac_paired"] for r in ob])
    ob_sld = np.array([r["sld_per100nt"] for r in ob])
    ob_ml = np.array([r["max_loop"] for r in ob])

    pcts = {
        "obelisk_pct_with_frac_paired_<=_6S": float(100 * (ob_fp <= fp).mean()),
        "obelisk_pct_with_sld_<=_6S": float(100 * (ob_sld <= sld).mean()),
        "obelisk_pct_with_max_loop_>=_6S": float(100 * (ob_ml >= ml).mean()),
    }
    print("\n6S RNA quantile in obelisk catalog:")
    print(f"  fp  = {fp:.3f}: {pcts['obelisk_pct_with_frac_paired_<=_6S']:.1f}% below")
    print(f"  sld = {sld:.2f}:  {pcts['obelisk_pct_with_sld_<=_6S']:.1f}% below")
    print(f"  max = {ml}:    {pcts['obelisk_pct_with_max_loop_>=_6S']:.1f}% above")

    payload = {
        "tool": "sixs_comparison",
        "sequence_id": "E.coli K-12 ssrS (6S RNA, 184 nt)",
        "length_nt": len(SIXS_ECOLI),
        "MFE_kcal_per_mol": float(mfe),
        "frac_paired": float(fp),
        "sld_per100nt": float(sld),
        "max_loop_nt": int(ml),
        "BRRC_pass_per_criterion": {k: bool(v) for k, v in passes.items()},
        "ob_catalog_q": pcts,
    }
    out = JSON / "sixs_comparison.json"
    out.write_text(json.dumps(payload, indent=2))
    print(f"\nwrote {out}")
    return payload

rnap_CMDS = {"promoter": promoter, "sixs": sixs}

def rnap_main(argv=None):
    argv = sys.argv[1:] if argv is None else argv
    cmd = argv[0] if argv else "promoter"
    if cmd not in rnap_CMDS:
        print(rnap_HELP, file=sys.stderr); sys.exit(2)
    rnap_CMDS[cmd]()

if __name__ == "__main__":
    rnap_main()

stats_HELP = "stats: cross | cluster | enrichment | run"

def _wilson_ci(k, n, alpha=0.05):
    if n == 0:
        return 0.0, 0.0
    z = norm.ppf(1 - alpha / 2)
    phat = k / n
    denom = 1 + z * z / n
    centre = (phat + z * z / (2 * n)) / denom
    half = (z * math.sqrt(phat * (1 - phat) / n + z * z / (4 * n * n))) / denom
    return max(0.0, centre - half), min(1.0, centre + half)

def _cliffs_delta(a, b):
    a = np.asarray(a); b = np.asarray(b)
    if not len(a) or not len(b):
        return 0.0
    n_gt = sum(1 for x in a for y in b if x > y)
    n_lt = sum(1 for x in a for y in b if x < y)
    return (n_gt - n_lt) / (len(a) * len(b))

CROSS_METRICS = ["mfe_per_nt", "frac_paired",
                 "sld_per100nt", "max_loop"]

def cross():
    m = json.load(open(JSON / "master_features_expanded.json"))
    groups = {"OBELISK": m["OBELISK_n27"],
              "DELTA":   m["DELTA_n25"],
              "RFAM":    m["RFAM_EXPANDED_n750"]}
    pairs = [("OBELISK", "DELTA"), ("OBELISK", "RFAM"), ("DELTA", "RFAM")]
    n_tests = len(CROSS_METRICS) * len(pairs)
    bonf_alpha = 0.05 / n_tests
    rows = []
    for metric in CROSS_METRICS:
        for a, b in pairs:
            xa = [r[metric] for r in groups[a]]
            xb = [r[metric] for r in groups[b]]
            u, p = mannwhitneyu(xa, xb, alternative="two-sided")
            d = _cliffs_delta(xa, xb)
            sig = p < bonf_alpha
            rows.append({"metric": metric, "comparison": f"{a}_vs_{b}",
                         "n_a": len(xa), "n_b": len(xb),
                         "U": float(u),
                         "p_raw": float(p),
                         "cliffs_delta": round(float(d), 4),
                         "passes_bonferroni": bool(sig)})
    out = {"tool": "cross_class_stats",
           "n_tests": n_tests,
           "bonferroni_alpha": bonf_alpha,
           "rows": rows}
    p = JSON / "cross_class_stats.json"
    p.write_text(json.dumps(out, indent=2))
    print(f"wrote {p}  (n_tests={n_tests}, alpha_bonf={bonf_alpha:.5f})")
    return out

def cluster(n_boot=1000, seed=42):
    summary = json.load(open(JSON / "brrc_envelope_summary.json"))

    pc = {k: v for k, v in summary["per_cluster"].items() if v["n"] >= 1}
    rng = np.random.default_rng(seed)
    names = sorted(pc.keys())
    rates = np.array([100 * pc[k]["all_three_pass"] / pc[k]["n"]
                      for k in names])
    weights = np.array([pc[k]["n"] for k in names], dtype=float)
    obs_eq = float(rates.mean())
    obs_w = float(np.average(rates, weights=weights))

    boot_eq, boot_w = [], []
    n_clusters = len(names)
    for _ in range(n_boot):
        idx = rng.integers(0, n_clusters, n_clusters)
        boot_eq.append(float(rates[idx].mean()))
        boot_w.append(float(np.average(rates[idx], weights=weights[idx])))
    boot_eq = np.asarray(boot_eq); boot_w = np.asarray(boot_w)

    loo_eq = []
    for i in range(n_clusters):
        keep = [j for j in range(n_clusters) if j != i]
        loo_eq.append({"left_out": names[i],
                       "equal_weight_pct": float(rates[keep].mean())})
    centroid_only_pct = float(rates.mean())

    out = {"tool": "cluster_level_bootstrap",
           "n_clusters": n_clusters,
           "n_bootstrap_iterations": n_boot,
           "obs_eq_weight_pct": round(obs_eq, 2),
           "obs_n_wt_pct": round(obs_w, 2),
           "equal_weight_bootstrap_95_CI": [round(float(np.percentile(boot_eq, 2.5)), 2),
                                            round(float(np.percentile(boot_eq, 97.5)), 2)],
           "n_weighted_bootstrap_95_CI": [round(float(np.percentile(boot_w, 2.5)), 2),
                                          round(float(np.percentile(boot_w, 97.5)), 2)],
           "loco_min_pct": round(float(min(d["equal_weight_pct"] for d in loo_eq)), 2),
           "loco_max_pct": round(float(max(d["equal_weight_pct"] for d in loo_eq)), 2),
           "loco_chi_excl_pct": round(
               float(next(d["equal_weight_pct"] for d in loo_eq if d["left_out"] == "chi")), 2)
               if any(d["left_out"] == "chi" for d in loo_eq) else None,
           "leave_one_cluster_out": loo_eq,
           "centroid_only_pct": round(centroid_only_pct, 2),
           }
    p = JSON / "cluster_level_bootstrap.json"
    p.write_text(json.dumps(out, indent=2))
    print(f"wrote {p}")
    print(f"  equal-weight: {obs_eq:.2f}%, "
          f"95% CI {out['equal_weight_bootstrap_95_CI']}")
    return out

def enrichment():
    s = json.load(open(JSON / "brrc_envelope_summary.json"))
    ko = s["OBELISK"]["all_three_pass"]; no = s["OBELISK"]["n"]
    ks = s["SHUFFLED"]["all_three_pass"]; ns = s["SHUFFLED"]["n"]
    ob_lo, ob_hi = _wilson_ci(ko, no)
    sh_lo, sh_hi = _wilson_ci(ks, ns)
    ob_pt = ko / no; sh_pt = ks / ns

    if ks > 0 and sh_pt > 0:
        ratio_pt = ob_pt / sh_pt

        se_log = math.sqrt((1 - ob_pt) / ko + (1 - sh_pt) / ks)
        z = 1.96
        log_rr = math.log(ratio_pt)
        katz_lo = math.exp(log_rr - z * se_log)
        katz_hi = math.exp(log_rr + z * se_log)

        bound_lo = ob_lo / sh_hi if sh_hi > 0 else None
        bound_hi = ob_hi / sh_lo if sh_lo > 0 else None
    else:
        ratio_pt = None
        katz_lo = katz_hi = None
        bound_lo = bound_hi = None

    out = {"tool": "enrichment_ci",
           "obelisk": {"k": ko, "n": no, "pct": round(100 * ob_pt, 2),
                       "wilson_95_CI_pct": [round(100 * ob_lo, 2),
                                            round(100 * ob_hi, 2)]},
           "shuffle": {"k": ks, "n": ns, "pct": round(100 * sh_pt, 4),
                       "wilson_95_CI_pct": [round(100 * sh_lo, 4),
                                            round(100 * sh_hi, 4)]},
           "fold_enrichment_point": (round(ratio_pt, 1)
                                     if ratio_pt is not None else None),
           "fold_enrichment_Katz_logRR_95_CI": [round(katz_lo, 0),
                                                round(katz_hi, 0)]
               if katz_lo is not None else None,
           "fold_enrich_cons": [round(bound_lo, 0),
                                                  round(bound_hi, 0)]
               if bound_lo is not None else None,
           "rule_of_three_pct":
               round(100 * 3 / ns, 4),
           "method": "Wilson + Katz log-RR"}
    p = JSON / "enrichment_ci.json"
    p.write_text(json.dumps(out, indent=2))
    print(f"wrote {p}")
    print(f"  obelisk: {out['obelisk']['pct']}% "
          f"95% CI {out['obelisk']['wilson_95_CI_pct']}")
    print(f"  shuffle: {out['shuffle']['pct']}% "
          f"95% CI {out['shuffle']['wilson_95_CI_pct']}")
    print(f"  enrichment point: {out['fold_enrichment_point']}x")
    print(f"  Katz log-RR 95% CI: {out['fold_enrichment_Katz_logRR_95_CI']}")
    print(f"  Conservative bound: {out['fold_enrich_cons']}")
    return out

def run():
    cross()
    cluster()
    enrichment()

stats_CMDS = {"cross": cross, "cluster": cluster,
        "enrichment": enrichment, "run": run}

def stats_main(argv=None):
    argv = sys.argv[1:] if argv is None else argv
    cmd = argv[0] if argv else "run"
    if cmd not in stats_CMDS:
        print(stats_HELP, file=sys.stderr); sys.exit(2)
    stats_CMDS[cmd]()

if __name__ == "__main__":
    stats_main()

def _load_catalog(xlsx):
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    ws = wb["Page 1"]
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or row[0] is None:
            if i > 0 and row[0] is None:
                break
            continue
        seq = row[6]
        if seq and isinstance(seq, str) and len(seq) > 200:
            out.append({"id": str(row[0]), "name": row[1],
                        "seq": seq.upper().replace("T", "U")})
    return out

def _write_sto(records, ss_cons, out_path, name):
    L = len(records[0]["seq"])
    pad = max(max(len(r["id"]) for r in records), len("#=GC SS_cons")) + 2
    with open(out_path, "w") as fh:
        fh.write("# STOCKHOLM 1.0\n")
        fh.write(f"#=GF ID  obelisk_{name}\n")
        fh.write(f"#=GF DE  Obelisk subfamily {name} same-length MSA "
                 f"(n={len(records)}, L={L}); SS_cons from RNAfold MFE.\n\n")
        for r in records:
            fh.write(f"{r['id']:<{pad}}{r['seq']}\n")
        fh.write(f"{'#=GC SS_cons':<{pad}}{ss_cons}\n//\n")

def dump():
    import RNA
    cat = _load_catalog(RAW / "zheludev_mmc2.xlsx")
    by = defaultdict(list)
    for c in cat:
        if c["name"] and c["name"] != ".":
            by[c["name"]].append(c)
    print(f"loaded {len(cat)} obelisks; {len(by)} named clusters")
    for target in ("alpha", "omega"):
        recs = by.get(target, [])
        if not recs:
            print(f"  {target}: missing"); continue
        L_cnt = Counter(len(r["seq"]) for r in recs)
        common_L, _ = L_cnt.most_common(1)[0]
        same = [r for r in recs if len(r["seq"]) == common_L]
        if len(same) < 10:
            print(f"  {target}: too few same-length"); continue
        db, mfe = RNA.fold(same[0]["seq"])
        out = PROC / f"obelisk_{target}_subfamily_msa.sto"
        _write_sto(same, db, out, target)
        print(f"  {target}: n={len(same)}, L={common_L}, "
              f"pairs={db.count('(')}, MFE={mfe:.1f}  ->  {out.name}")

def _parse_one(name):
    cov = (RSCAPE / f"obelisk_{name}.cov").read_text()
    log = (RSCAPE / f"{name}.log").read_text()
    power = (RSCAPE / f"obelisk_{name}.power").read_text().splitlines()
    out = {"cluster": name}
    m = re.search(r"nseq\s+(\d+).*alen\s+(\d+).*avgid\s+([\d.]+).*nbpairs\s+(\d+)",
                  log)
    if m:
        out.update(n_seqs=int(m.group(1)), alignment_length=int(m.group(2)),
                   avg_pairwise_identity_pct=float(m.group(3)),
                   n_basepairs_in_input_structure=int(m.group(4)))
    m = re.search(
        r"BPAIRS\s+(\d+)\s*\n# avg substitutions per BP\s+([\d.]+)\s*\n"
        r"# BPAIRS expected to covary\s+([\d.]+)\s*\+/-\s*([\d.]+)\s*\n"
        r"# BPAIRS observed to covary\s+(\d+)", cov)
    if m:
        out.update(rscape_n_basepairs=int(m.group(1)),
                   avg_substitutions_per_basepair=float(m.group(2)),
                   expected_covarying_under_null=float(m.group(3)),
                   expected_covarying_se=float(m.group(4)),
                   observed_covarying=int(m.group(5)))
    m = re.search(r"Method Target_E-val.*\n# (\S+)\s+([\d.]+)\s+"
                  r"\[([-\d.]+),([-\d.]+)\]\s+\[(\d+) \| (\d+) (\d+) (\d+) "
                  r"\| ([\d.]+) ([\d.]+) ([\d.]+)\]", cov)
    if m:
        out.update(method=m.group(1), target_e_value=float(m.group(2)),
                   fp_significant=int(m.group(5)),
                   tp_significant=int(m.group(6)),
                   true_basepairs=int(m.group(7)),
                   found_significant=int(m.group(8)),
                   sensitivity=float(m.group(9)), ppv=float(m.group(10)),
                   f_score=float(m.group(11)))
    powered = []
    for line in power:
        if line.startswith("#") or not line.strip():
            continue
        parts = line.split()
        if len(parts) >= 4:
            try:
                powered.append(float(parts[3]))
            except ValueError:
                pass
    if powered:
        out.update(n_basepairs_with_power_data=len(powered),
                   mean_power_per_basepair=sum(powered) / len(powered),
                   max_power_per_basepair=max(powered),
                   n_basepairs_with_power_above_5pct=sum(
                       1 for p in powered if p > 0.05))
    return out

def parse():
    res = {
        "tool": "R-scape",
        "tool_release": "2.5.9",
        "subfamilies": [_parse_one("alpha"), _parse_one("omega")],
    }
    out = JSON / "rscape_official.json"
    out.write_text(json.dumps(res, indent=2))
    print(f"wrote {out}")
    return res

def rscape_main(argv=None):
    argv = sys.argv[1:] if argv is None else argv
    cmd = argv[0] if argv else "parse"
    if cmd == "dump":
        dump()
    elif cmd == "parse":
        parse()
    else:
        print(f"unknown command: {cmd!r}; use 'dump' or 'parse'",
              file=sys.stderr); sys.exit(2)

if __name__ == "__main__":
    rscape_main()

structure_prediction_HELP = """Usage:
  python -m oblin predict bundle
  python -m oblin predict af3 [model_index]
  python -m oblin predict boltz <boltz_out_dir>
  python -m oblin predict foldseek [pdb]
  python -m oblin predict compare
"""

OBLIN1 = ("MRDIELDSSAFRSQVSLLSQETSEKFLTGAALVSPKRSKYYISEVEGLKVHSRSKKDLLA"
          "LAIISWWLEDSIRFYLQEELYFLSLNNSDLIEIRLCLTSKSGMLNFLEDTTLYHSRDLFG"
          "NILPTSPEKQVRLANLVSVRYGPTSLPKRVIRRRGYKDHGSRRFPHEVHDLSSGKLAQIK"
          "YEEEIQSYHDTLLFLRGWLDGF")
DOM_A = (130, 175)
METAL = {158, 159, 166}

AA3 = {"A": "ALA", "R": "ARG", "N": "ASN", "D": "ASP", "C": "CYS",
       "E": "GLU", "Q": "GLN", "G": "GLY", "H": "HIS", "I": "ILE",
       "L": "LEU", "K": "LYS", "M": "MET", "F": "PHE", "P": "PRO",
       "S": "SER", "T": "THR", "W": "TRP", "Y": "TYR", "V": "VAL"}

def _ca_dist(ca, a, b):
    if a not in ca or b not in ca:
        return None
    return math.sqrt(sum((ca[a][i] - ca[b][i]) ** 2 for i in range(3)))

def _rel(path):
    path = Path(path)
    try:
        return path.resolve().relative_to(ROOT).as_posix()
    except ValueError:
        return path.as_posix()

_FASTA = (
    ">oblin1 Streptococcus_sanguinis_obelisk-S.s_centroid Oblin-1_protein 202aa\n"
    + "\n".join(OBLIN1[i:i+60] for i in range(0, len(OBLIN1), 60)) + "\n"
)

def bundle():
    out = PROC / "oblin1_af3_input"
    out.mkdir(parents=True, exist_ok=True)
    src = PROC / "oblin1_colabfold" / "bfd.mgnify30.metaeuk30.smag30.a3m"
    if src.exists():
        shutil.copy(src, out / "oblin1_environmental.a3m")
    (out / "oblin1.fasta").write_text(_FASTA)
    print(f"AF3 bundle written to {out}")

def _parse_cif_ca(cif):
    out = {}
    cols, in_loop = [], False
    parse_errors = 0
    with open(cif) as fh:
        for line in fh:
            line = line.rstrip()
            if line.startswith("loop_"):
                in_loop = True; cols = []
                continue
            if in_loop and line.startswith("_atom_site."):
                cols.append(line.split(".", 1)[1].strip()); continue
            if in_loop and (line.startswith("ATOM") or line.startswith("HETATM")):
                tok = line.split()
                rec = dict(zip(cols, tok))
                if rec.get("label_atom_id") == "CA" or rec.get("auth_atom_id") == "CA":
                    try:
                        rn = int(rec.get("auth_seq_id", rec.get("label_seq_id")))
                        x = float(rec["Cartn_x"]); y = float(rec["Cartn_y"])
                        z = float(rec["Cartn_z"])
                        b = float(rec.get("B_iso_or_equiv", 0))
                        out[rn] = (x, y, z, b)
                    except (KeyError, ValueError):
                        parse_errors += 1
            elif in_loop and line.startswith("#"):
                in_loop = False; cols = []
    if not out:
        raise ValueError(
            f"no CA atoms parsed from {cif} ({parse_errors} CA records "
            f"hit KeyError/ValueError); refusing to silently emit pLDDT=0")
    return out

def _ca_to_pdb(ca, seq, out_path):
    with open(out_path, "w") as fh:
        for k, rn in enumerate(sorted(ca.keys()), 1):
            x, y, z, b = ca[rn]
            res = AA3.get(seq[rn-1], "UNK") if rn <= len(seq) else "UNK"
            fh.write(f"ATOM  {k:5d}  CA  {res} A{rn:4d}    "
                     f"{x:8.3f}{y:8.3f}{z:8.3f}  1.00{b:6.2f}           C\n")
        fh.write("END\n")

def af3(idx=0):
    d = RESULTS / "af3"
    cifs = sorted(d.glob("fold_*_model_*.cif"))
    sums = sorted(d.glob("fold_*_summary_confidences_*.json"))
    if not cifs:
        print(f"no AF3 CIF in {d}", file=sys.stderr); sys.exit(1)
    if idx >= len(cifs):
        print(f"only {len(cifs)} models, asked for {idx}", file=sys.stderr)
        sys.exit(1)
    cif = cifs[idx]
    summ = sums[idx] if idx < len(sums) else None
    print(f"reading {cif.name}")
    ca = _parse_cif_ca(cif)
    print(f"  {len(ca)} CA atoms")

    plddt = [ca[r][3] for r in sorted(ca)]
    mean = sum(plddt) / len(plddt) if plddt else 0
    domA = [ca[r][3] for r in sorted(ca) if DOM_A[0] <= r <= DOM_A[1]]
    domA_mean = sum(domA) / len(domA) if domA else 0

    out = {
        "tool": "AlphaFold3 Server",
        "model_index": idx,
        "cif": _rel(cif),
        "n_residues_with_CA": len(ca),
        "mean_CA_pLDDT": round(mean, 2),
        "domainA_mean_CA_pLDDT": round(domA_mean, 2),
        "n_residues_above_pLDDT_70": sum(1 for v in plddt if v >= 70),
        "n_residues_above_pLDDT_90": sum(1 for v in plddt if v >= 90),
        "metal_cluster_CA_distances_A": {
            "D158_H159": round(_ca_dist(ca, 158, 159) or 0, 2),
            "D158_H166": round(_ca_dist(ca, 158, 166) or 0, 2),
            "H159_H166": round(_ca_dist(ca, 159, 166) or 0, 2),
        },
    }
    if summ:
        try:
            out["summary_confidences"] = json.loads(summ.read_text())
        except Exception as exc:
            out["summary_confidences_error"] = str(exc)

    pdb = PROC / "oblin1_af3.pdb"
    _ca_to_pdb(ca, OBLIN1, pdb)
    j = JSON / "oblin1_af3.json"
    j.write_text(json.dumps(out, indent=2))
    print(f"\nwrote {pdb}\nwrote {j}")
    print(f"  mean pLDDT = {mean:.2f}  Domain-A = {domA_mean:.2f}")
    return out
def _parse_pdb_ca(pdb):
    out = {}
    with open(pdb) as fh:
        for line in fh:
            if not line.startswith("ATOM") or line[12:16].strip() != "CA":
                continue
            rn = int(line[22:26])
            x = float(line[30:38]); y = float(line[38:46]); z = float(line[46:54])
            b = float(line[60:66])
            out[rn] = (x, y, z, b)
    return out

def boltz(out_dir=None):
    if out_dir is None:
        if len(sys.argv) < 3:
            print("usage: predict boltz <boltz_out_dir>", file=sys.stderr)
            sys.exit(2)
        out_dir = sys.argv[2]
    out_dir = Path(out_dir)
    pdbs = sorted(out_dir.rglob("*.pdb"))
    confs = sorted(out_dir.rglob("*confidence*.json"))
    if not pdbs:
        print(f"no PDBs under {out_dir}", file=sys.stderr); sys.exit(1)

    pdb = pdbs[0]
    conf = confs[0] if confs else None
    print(f"reading {pdb}")
    ca = _parse_pdb_ca(pdb)
    plddt = [ca[r][3] for r in sorted(ca)]
    mean = sum(plddt) / len(plddt) if plddt else 0
    domA = [ca[r][3] for r in sorted(ca) if DOM_A[0] <= r <= DOM_A[1]]
    domA_mean = sum(domA) / len(domA) if domA else 0

    out = {
        "tool": "Boltz-2",
        "input_pdb": _rel(pdb),
        "n_residues": len(ca),
        "mean_pLDDT": round(mean, 2),
        "domainA_mean_pLDDT": round(domA_mean, 2),
        "n_residues_above_pLDDT_70": sum(1 for v in plddt if v >= 70),
        "n_residues_above_pLDDT_90": sum(1 for v in plddt if v >= 90),
        "metal_cluster_CA_distances_A": {
            "D158_H159": round(_ca_dist(ca, 158, 159) or 0, 2),
            "D158_H166": round(_ca_dist(ca, 158, 166) or 0, 2),
            "H159_H166": round(_ca_dist(ca, 159, 166) or 0, 2),
        },
    }
    if conf:
        try:
            out["boltz_confidence_json"] = json.loads(conf.read_text())
        except Exception as exc:
            out["boltz_confidence_json_error"] = str(exc)

    pdb_dest = PROC / "oblin1_boltz2.pdb"
    pdb_dest.write_text(pdb.read_text())
    j = JSON / "oblin1_boltz2.json"
    j.write_text(json.dumps(out, indent=2))
    print(f"\nwrote {pdb_dest}\nwrote {j}")
    print(f"  mean pLDDT = {mean:.2f}  Domain-A = {domA_mean:.2f}")
    return out
_FOLDSEEK = "https://search.foldseek.com/api"
_DBS = ["pdb100", "afdb50", "gmgcl_id", "mgnify_esm30", "cath50"]

def foldseek(pdb_name="oblin1_boltz2.pdb"):
    pdb = PROC / pdb_name
    print(f"submitting {pdb} to foldseek (3diaa, {len(_DBS)} dbs)")

    with open(pdb, "rb") as fh:
        files = {"q": (pdb.name, fh.read(), "application/octet-stream")}
        data = [("database[]", db) for db in _DBS] + [("mode", "3diaa")]
        r = requests.post(f"{_FOLDSEEK}/ticket", files=files, data=data,
                          timeout=90)
    r.raise_for_status()
    ticket = r.json()["id"]
    print(f"  ticket = {ticket}")

    deadline = time.time() + 600
    while time.time() < deadline:
        rs = requests.get(f"{_FOLDSEEK}/ticket/{ticket}", timeout=30)
        rs.raise_for_status()
        s = rs.json().get("status")
        if s == "COMPLETE":
            break
        if s == "ERROR":
            raise RuntimeError(f"foldseek ERROR for {ticket}")
        time.sleep(6)
    else:
        raise TimeoutError("foldseek did not finish within 600s")

    payload = requests.get(f"{_FOLDSEEK}/result/{ticket}/0",
                           timeout=60).json()
    summary = {"tool": "foldseek_3diaa_on_boltz2", "ticket": ticket,
               "databases": [], "per_db": {}}
    for r in payload.get("results", []):
        db = r["db"]
        summary["databases"].append(db)
        hits = []
        for grp in r.get("alignments", []) or []:
            if grp:
                hits.extend(grp)
        hits.sort(key=lambda h: float(h.get("eval", 99)))
        summary["per_db"][db] = {
            "n_hits_total": len(hits),
            "n_hits_significant_e_lt_0.05": sum(
                1 for h in hits if float(h.get("eval", 99)) < 0.05),
            "top_hit": ({k: v for k, v in hits[0].items()
                         if k not in ("tCa", "qAln", "dbAln")}
                        if hits else None),
        }
    tag = pdb_name.replace("oblin1_", "").replace(".pdb", "")
    out = JSON / f"oblin1_foldseek_{tag}.json"
    out.write_text(json.dumps(summary, indent=2))
    print(f"\nwrote {out}")
    for db, info in summary["per_db"].items():
        n, ns = info["n_hits_total"], info["n_hits_significant_e_lt_0.05"]
        top = info["top_hit"]
        e = float(top["eval"]) if top else None
        print(f"  {db:14s}  {n:3d} hits ({ns} sig)" +
              (f"  top e={e:.2g}" if e is not None else ""))
    return summary
def _load_ca(p):
    out = {}
    with open(p) as fh:
        for line in fh:
            if not line.startswith("ATOM") or line[12:16].strip() != "CA":
                continue
            rn = int(line[22:26])
            x = float(line[30:38]); y = float(line[38:46]); z = float(line[46:54])
            out[rn] = np.array([x, y, z])
    return out

def _kabsch(P, Q):
    if len(P) < 3 or len(Q) < 3:
        raise ValueError(f"_kabsch needs >=3 atom pairs; got {len(P)}/{len(Q)}")
    Pc, Qc = P.mean(axis=0), Q.mean(axis=0)
    H = (P - Pc).T @ (Q - Qc)
    U, _, Vt = np.linalg.svd(H)
    d = np.sign(np.linalg.det(Vt.T @ U.T))
    R = Vt.T @ np.diag([1, 1, d]) @ U.T
    return R, Qc - R @ Pc

def _tm(d, L):
    if L == 0 or len(d) == 0:
        return None
    if L >= 21:
        d0 = max(0.5, 1.24 * (L - 15) ** (1/3) - 1.8)
    else:
        d0 = 0.5
    return float(np.mean(1.0 / (1.0 + (d / d0) ** 2)))

def compare():
    af = _load_ca(PROC / "oblin1_af3.pdb")
    bo = _load_ca(PROC / "oblin1_boltz2.pdb")
    common = sorted(set(af) & set(bo))
    print(f"AF3 {len(af)}; Boltz {len(bo)}; common {len(common)}")

    P = np.array([bo[r] for r in common])
    Q = np.array([af[r] for r in common])

    R, t = _kabsch(P, Q)
    deltas = np.linalg.norm((R @ P.T).T + t - Q, axis=1)
    rmsd = float(np.sqrt((deltas ** 2).mean()))
    tm = _tm(deltas, len(Q))

    n_idx = [i for i, r in enumerate(common) if 1 <= r <= 130]
    a_idx = [i for i, r in enumerate(common) if 130 < r <= 175]
    rmsd_n = rmsd_da_in_n = None
    if n_idx:
        Pn, Qn = P[n_idx], Q[n_idx]
        Rn, tn = _kabsch(Pn, Qn)
        d_n = np.linalg.norm((Rn @ P.T).T + tn - Q, axis=1)
        rmsd_n = float(np.sqrt((d_n[n_idx] ** 2).mean()))
        if a_idx:
            rmsd_da_in_n = float(np.sqrt((d_n[a_idx] ** 2).mean()))

    def regional(lo_inc, hi_inc, exclude_lo=False):
        if exclude_lo:
            idx = [i for i, r in enumerate(common) if lo_inc < r <= hi_inc]
        else:
            idx = [i for i, r in enumerate(common) if lo_inc <= r <= hi_inc]
        if not idx:
            return None
        return float(np.sqrt((deltas[idx] ** 2).mean()))

    out = {
        "tool": "structure_compare",
        "compared": "AlphaFold3 vs Boltz-2 (common CA atoms)",
        "n_residues_common": len(common),
        "rmsd_A_after_global_Kabsch": round(rmsd, 3),
        "TM_score_approx": round(tm, 3),
        "rmsd_A_N_globule_aligned_locally": (round(rmsd_n, 3)
                                             if rmsd_n is not None else None),
        "rmsd_A_Domain_A_in_N_globule_frame": (round(rmsd_da_in_n, 3)
                                                if rmsd_da_in_n is not None else None),
        "regional_RMSD_A": {
            "N_globule_1_130": regional(1, 130),
            "Domain_A_131_175": regional(130, 175, exclude_lo=True),
            "C_helix_176_202": regional(175, 202, exclude_lo=True),
        },
    }
    out_path = JSON / "structure_compare.json"
    out_path.write_text(json.dumps(out, indent=2))
    print(f"\nwrote {out_path}")
    print(f"  global Kabsch RMSD = {rmsd:.2f} A")
    print(f"  TM-score approx    = {tm:.2f}")
    if rmsd_n is not None:
        print(f"  N-globule RMSD     = {rmsd_n:.2f} A")
    if rmsd_da_in_n is not None:
        print(f"  Domain-A in N-frame = {rmsd_da_in_n:.2f} A")
    return out

def predict_main(argv=None):
    argv = sys.argv[1:] if argv is None else argv
    cmd = argv[0] if argv else "bundle"
    rest = argv[1:]
    if cmd == "bundle":
        bundle()
    elif cmd == "af3":
        af3(int(rest[0]) if rest else 0)
    elif cmd == "boltz":
        boltz(rest[0] if rest else None)
    elif cmd == "foldseek":
        foldseek(rest[0] if rest else None)
    elif cmd == "compare":
        compare()
    else:
        print(structure_prediction_HELP, file=sys.stderr); sys.exit(2)

if __name__ == "__main__":
    predict_main()
